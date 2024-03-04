import streamlit as st
from streamlit_option_menu import option_menu

with st.sidebar:
    st.write("This is a demo software program with some demo modules. "
             "The purpose of the demo is to show the data processing capability of our applications.  \n"
             "  \n"
             "If you need any tailor made software with similar data processing features please contact us.  \n"
             "  \n"
             "")
    selected = option_menu("RAN Modules",[
        "CDR Reporting",
        "CDR Reports Historical Comparison",
        "View Historical Reports",
        "View on Map",
        # "path layer"
        ], menu_icon="cast", default_index=0)

if selected == "CDR Reporting":
    import pandas as pd
    import webbrowser
    import numpy as np
    import zipfile

    st.title('CDR Reporting')
    st.header('Drive Test Statistical Report Module - Data')
    st.write("This module is a demo for CDR Statistical Reporting. ")


    uploaded_file = st.file_uploader("Choose input zip file to process", type="zip", accept_multiple_files=True)

    if st.button('Process uploaded files'):

        with zipfile.ZipFile(uploaded_file[0], 'r') as zip_ref:
            csv_file_name = zip_ref.namelist()[0]
            zip_ref.extract(csv_file_name)


        # Script start

        # path = "C:/08_2024_DT/HiDrive-CDR/Master Data/O2_CDR_DATA_RATINGEN_20240207/"
        # csv_file_name = path + "Metric Group 1.csv"

        df = pd.read_csv(csv_file_name, low_memory=False).head(100000)

        sel_columns = ["Time", "Date", "Latitude", "Longitude", "Grouping", "Grouping_with_Direction", "Grouping_with_Direction_HTTP", "Grouping_with_Direction_Ping", "Operator", "Task_Type", "Technology", "HTTP_URL", "Streaming_URL", "Ping_Address", "Serving Cell RS SINR (dB)", "Serving Cell RSRP (dBm)", "Serving Cell RSRQ (dB)", "HTTP_Download_Average_Throughput", "HTTP_Download_Service_Average_Throughput", "HTTP_Download_Session_Failure_Ratio", "HTTP_Download_Session_Success_Ratio",
            "HTTP_Outcome", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Download_Data_Transfer_Success_Ratio_Method_A", "HTTP_Download_Data_Transfer_Time_sec_Method_A", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Time_sec_Method_A", "HTTP_Download_Mean_Data_Rate_kbps_Method_A", "HTTP_Download_Transfer_Start_Delay_Method_A", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B",
            "HTTP_Download_Data_Transfer_Success_Ratio_Method_B", "HTTP_Download_Data_Transfer_Time_sec_Method_B", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Time_sec_Method_B", "HTTP_Download_Mean_Data_Rate_kbps_Method_B", "HTTP_Upload_Average_Throughput", "HTTP_Upload_Service_Average_Throughput", "HTTP_Upload_Service_Transfer_Time", "HTTP_Upload_Session_Failure_Ratio",
            "HTTP_Upload_Session_Success_Ratio", "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Time_sec_Method_A", "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A", "HTTP_Upload_Mean_Data_Rate_kbps_Method_A", "HTTP Transfer Time (second)", "Streaming_Outcome_Type", "Aborted_by_User",
            "Streaming_Average_Session_Resolution", "Streaming_Average_Throughput", "Streaming_Completion_Rate", "Streaming_Duration", "Streaming_HD_Resolution", "Streaming_HD_Resolution_Ratio", "Streaming_Impairment_Free", "Streaming_Impairment_Free_Video_Session_Ratio", "Streaming_Maximum_Duration_Of_Video_Session_Interruptions", "Streaming_Number_Of_Video_Session_Interruptions", "Streaming_Player_Size_kB", "Streaming_Reproduction_Cutoff_Ratio", "Streaming_Reproduction_Start_Delay_sec",
            "Streaming_Reproduction_Start_Failure_Ratio", "Streaming_Service_Access_Time_ms", "Streaming_Service_Access_Time_sec", "Streaming_Session_Failure_Ratio", "Streaming_Session_Qualified", "Streaming_Session_Qualified_Ratio", "Streaming_Session_Video_Interruption_Duration", "Streaming_Session_Without_Interruption_Rate", "Streaming_Setup_Success_Rate", "Streaming_State_Prebuffering_to_Reproducing_Delay", "Streaming_State_Request_to_Prebuffering_Delay",
            "Streaming_State_Request_to_Reproducing_Delay", "Streaming_Success_Rate", "Streaming_Throughput_Filtered", "Streaming_Total_Duration_Of_Video_Session_Interruptions", "Streaming_Video_Buffer_Size_kB", "Streaming_Video_IP_Service_Access_Time_ms", "Streaming_Video_IP_Service_Access_Time_sec", "Streaming_Video_Play_Start_Failure_Ratio", "Streaming_Video_Play_Start_Time_sec", "Streaming_Video_Session_Cutoff_Ratio", "Streaming_Video_Session_Failure_Ratio",
            "Streaming_Video_Session_Success_Ratio", "Streaming_Video_Session_Time_sec", "Streaming_Video_Size_kB", # "Streaming_Impairment_Free",
            # "Streaming_Reproduction_Cutoff_Ratio",
            "SessionTime", "SessionTime_Upload", "ThroughputCountOver1MBit", "ThroughputCountOver3MBit", "ThroughputPercentageOver3MBit", "TCP_Handshake_Time_sec", "Ping_Count_Attempts", "Ping_Count_Failed", "Ping_Count_Success", "Ping_Delay_ms_Avg", "Ping_Delay_ms_Max", "Ping_Delay_ms_Min", "Ping_Packet_Loss_Rate", "Ping_Packet_Success_Rate", "Ping_Roundtrip_Time_ms", "Ping_Roundtrip_Time_sec", "Ping_Size", "Data_Radio_Bearer", "Fixed_Duration", "IP_Interruption_Time_ms", "Is_Multi_RAB",
            "LTE_Serving_Cell_Count_Average", "Radio_Access_Technology", "Service_Bearer", "FTP_Download_Average_Throughput_Not_Completed_Session", "FTP_Download_Bearer", "FTP_Download_Error_Cause", "FTP_Download_Outcome", "FTP_Download_Session_Failure_Ratio", "FTP_Download_Session_Success_Ratio", "FTP_Server_File", "Seconds_Start_to_End_FTP_Download", "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_A", "FTP_Download_Data_Transfer_Success_Ratio_Method_A",
            "FTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "FTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "FTP_Download_IP_Service_Setup_Time_sec_Method_A", "FTP_Download_Transfer_Start_Delay_Method_A", "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_B", "FTP_Download_Data_Transfer_Success_Ratio_Method_B", "FTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "FTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "FTP_Download_IP_Service_Setup_Time_sec_Method_B",
            "Seconds_Start_to_End_FTP_Upload", "FTP_Upload_Bearer", "FTP_Upload_Outcome", "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_A", "FTP_Upload_IP_Service_Setup_Time_sec_Method_A", "FTP_Upload_Transfer_Start_Delay_Method_A", "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_B", "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_B", "FTP_Upload_IP_Service_Setup_Time_sec_Method_B", "Seconds_Start_to_End_UDP_Download",
            "UDP_Download_Error_Cause", "Seconds_Start_to_End_UDP_Upload", "UDP_Upload_Error_Cause", "DNS_Client", "DNS_Domain_Name", "DNS_First_In_Session", "DNS_Host_Name_Resolution_Failure_Ratio", "DNS_Host_Name_Resolution_Time_sec", "DNS_Host_Name_Total_Resolution_Time_sec", "DNS_Resolved_Address", "DNS_Server_Address"]

        # for i in sel_columns:
        # print(i)

        df_sel = df[sel_columns]

        # Combining 3 columns to determinte test ID's
        columns_to_combine = ["HTTP_URL", "Streaming_URL", "Ping_Address"]

        # df_sel['CombinedColumn'] = df_sel[columns_to_combine].apply(lambda row: ' '.join(str(value) for value in row if pd.notna(value)), axis=1)
        df_sel.loc[:, 'CombinedColumn'] = df_sel[columns_to_combine].apply(lambda row: ' '.join(str(value) for value in row if pd.notna(value)), axis=1)
        df_sel.loc[df_sel['CombinedColumn'] == '', 'CombinedColumn'] = np.nan

        # Calculating Session ID's
        n_e_r = df_sel.index[df_sel['Ping_Address'].notna()].tolist() + [len(df_sel)]  # n_e_r = non empty rows on ping address
        labels = list(range(1, len(n_e_r) + 1))  # Numbering Session ID's
        df_sel['Session_ID'] = pd.cut(df_sel.index, bins=[0] + n_e_r, labels=labels, right=False)

        # Calculating Test ID's
        n_e_r_2 = df_sel.index[df_sel['CombinedColumn'].notna()].tolist() + [len(df_sel)]  # n_e_r_2 = non empty rows combined column
        labels_2 = list(range(1, len(n_e_r_2) + 1))  # Numbering test id's
        labels_3 = ["8.8.8.8"] + list(df_sel.loc[n_e_r_2[:-1], 'CombinedColumn'])  # First test comes from an empty row so 8.8.8.8 added
        df_sel['Test_ID'] = pd.cut(df_sel.index, bins=[0] + n_e_r_2, labels=labels_2, right=False)

        # Test Types are same as test numbering
        df_sel['Test_type'] = pd.cut(df_sel.index, bins=[0] + n_e_r_2, labels=labels_3, right=False, ordered=False)

        # Changing test type names
        my_dict = {"8.8.8.8": "PING", "URI: https://ash-speed.hetzner.com/1GB.bin": "FDTT http DL MT", "URI: http://d26kjwdsl72dzf.cloudfront.net/upload": "FDTT http UL MT", "URI: https://www.google.de/": "Google.de", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/Kepler_mobile/index.html": "Kepler", "URI: https://www.gmx.net/": "gmx.de", "URI: https://de.m.wikipedia.org/wiki/Europa": "Wikipedia", "URI: https://www.chip.de/": "Chip.de",
            "Source Uri: https://www.youtube.com/watch?v=BQwC_DJSdfE": "youtube", "URI: https://www.instagram.com/": "instagram", "URI: https://www.amazon.de/": "Amazon.de", "URI: http://212.183.159.230/5MB.zip": "FDFS http DL ST", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/upload/": "FDFS http UL ST", }
        df_sel['Test_type'] = df_sel['Test_type'].replace(my_dict)

        # region calculating general statistics

        Total_Stats = {}
        FDFS_DL_Attempts = df_sel['HTTP_Outcome'].count()
        Total_Stats.update({"FDFS_DL_Attempts": df_sel['HTTP_Outcome'].count()})
        FDFS_DL_Success = len(df_sel[df_sel['HTTP_Outcome'] == 'Service Status: Succeeded'])
        Total_Stats.update({"FDFS_DL_Success": FDFS_DL_Success})
        FDFS_DL_Failure = len(df_sel[df_sel['HTTP_Outcome'] == 'Service Status: Failed'])
        Total_Stats.update({"FDFS_DL_Failure": FDFS_DL_Failure})
        FDFS_DL_Success_Ratio = "{:.2%}".format(FDFS_DL_Success / FDFS_DL_Attempts)
        Total_Stats.update({"FDFS_DL_Success_Ratio": FDFS_DL_Success_Ratio})
        FDFS_UL_Attempts = df_sel['HTTP_Upload_Session_Success_Ratio'].count()
        Total_Stats.update({"FDFS_UL_Attempts": FDFS_UL_Attempts})
        FDFS_UL_Success = len(df_sel[df_sel['HTTP_Upload_Session_Success_Ratio'] == 100])
        Total_Stats.update({"FDFS_UL_Success": FDFS_UL_Success})
        FDFS_UL_Failure = len(df_sel[df_sel['HTTP_Upload_Session_Success_Ratio'] == 0])
        Total_Stats.update({"FDFS_UL_Failure": FDFS_UL_Failure})
        FDFS_UL_Success_Ratio = "{:.2%}".format(FDFS_UL_Success / FDFS_UL_Attempts)
        Total_Stats.update({"FDFS_UL_Success_Ratio": FDFS_UL_Success_Ratio})

        HTTP_Browsing_Attempts = df_sel['HTTP_Download_Session_Success_Ratio'].count()
        Total_Stats.update({'HTTP_Browsing_Attempts': HTTP_Browsing_Attempts})
        HTTP_Browsing_Access_Failure = len(df_sel[df_sel['HTTP_Download_Session_Success_Ratio'] == 0])
        Total_Stats.update({'HTTP_Browsing_Access_Failure': HTTP_Browsing_Access_Failure})
        HTTP_Browsing_Access_Success = len(df_sel[df_sel['HTTP_Download_Session_Success_Ratio'] == 100])
        Total_Stats.update({'HTTP_Browsing_Access_Success': HTTP_Browsing_Access_Success})
        HTTP_Browsing_Access_Success_Ratio = "{:.2%}".format(FDFS_DL_Success / FDFS_DL_Attempts)
        Total_Stats.update({'HTTP_Browsing_Access_Success': HTTP_Browsing_Access_Success})
        HTTP_Browsing_Min_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].min()
        Total_Stats.update({'HTTP_Browsing_Min_Transfer_Time_s': HTTP_Browsing_Min_Transfer_Time_s})
        HTTP_Browsing_10_PCTL_Transfer_Time_s = HTTP_Browsing_Min_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].quantile(0.01)
        Total_Stats.update({'HTTP_Browsing_10_PCTL_Transfer_Time_s': HTTP_Browsing_10_PCTL_Transfer_Time_s})
        HTTP_Browsing_Average_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].mean()
        Total_Stats.update({'HTTP_Browsing_Average_Transfer_Time_s': HTTP_Browsing_Average_Transfer_Time_s})
        HTTP_Browsing_Median_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].median()
        Total_Stats.update({'HTTP_Browsing_Median_Transfer_Time_s': HTTP_Browsing_Median_Transfer_Time_s})
        HTTP_Browsing_90_PCTL_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].quantile(0.9)
        Total_Stats.update({'HTTP_Browsing_90_PCTL_Transfer_Time_s': HTTP_Browsing_90_PCTL_Transfer_Time_s})
        HTTP_Browsing_Max_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].max()
        Total_Stats.update({'HTTP_Browsing_Max_Transfer_Time_s': HTTP_Browsing_Max_Transfer_Time_s})

        Video_Stream_Attempts = df_sel['Streaming_Session_Failure_Ratio'].count()
        Video_Stream_Failures = len(df_sel[df_sel['Streaming_Session_Failure_Ratio'] == 0])
        Video_Stream_Access_Time_sec = df_sel["Streaming_Service_Access_Time_sec"].mean()
        Video_Stream_Success_Ratio_PC = len(df_sel[df_sel['Streaming_Session_Failure_Ratio'] == 100])
        Video_Stream_Reproduction_Start_Delay_sec = df_sel["Streaming_State_Request_to_Prebuffering_Delay"].mean()
        Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay = df_sel["Streaming_State_Request_to_Reproducing_Delay"].mean()
        # Streaming_Impairment_Free = df_sel["Streaming_Impairment_Free"].mean()
        Streaming_Reproduction_Cut_off_Ratio = df_sel[("Streaming_Reproduction_Cutoff_Ratio")].mean()
        Streaming_Session_Without_Interruption_Rate_PC = df_sel["Streaming_Session_Without_Interruption_Rate"].mean()
        Streaming_Aggregated_Average_Session_Resolution_p = df_sel["Streaming_Average_Session_Resolution"].mean()
        Ping_Attempt = df_sel['Ping_Count_Attempts'].count()
        Ping_Success = len(df_sel[df_sel['Ping_Count_Success'] == 10])
        Ping_Failures = len(df_sel[df_sel['Ping_Count_Failed'] == 10])
        Ping_Success_rate = "{:.2%}".format(Ping_Success / Ping_Attempt)
        Ping_RTT_ms = df_sel["Ping_Roundtrip_Time_ms"].mean()
        Ping_Trace_Loss_PC = df_sel["Ping_Packet_Loss_Rate"].mean()

        Total_Stats.update({'Video_Stream_Attempts': Video_Stream_Attempts})
        Total_Stats.update({'Video_Stream_Failures': Video_Stream_Failures})
        Total_Stats.update({'Video_Stream_Access_Time_sec': Video_Stream_Access_Time_sec})
        Total_Stats.update({'Video_Stream_Success_Ratio_PC': Video_Stream_Success_Ratio_PC})
        Total_Stats.update({'Video_Stream_Reproduction_Start_Delay_sec': Video_Stream_Reproduction_Start_Delay_sec})
        Total_Stats.update({'Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay': Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay})
        # Total_Stats.update({'Streaming_Impairment_Free_Ratio_PC':Streaming_Impairment_Free})
        Total_Stats.update({'Streaming_Reproduction_Cut-off_Ratio': Streaming_Reproduction_Cut_off_Ratio})
        Total_Stats.update({'Streaming_Session_Without_Interruption_Rate_PC': Streaming_Session_Without_Interruption_Rate_PC})
        Total_Stats.update({'Streaming_Aggregated_Average_Session_Resolution_p': Streaming_Aggregated_Average_Session_Resolution_p})
        Total_Stats.update({'Ping_Attempt': Ping_Attempt})
        Total_Stats.update({'Ping_Success': Ping_Success})
        Total_Stats.update({'Ping_Failures': Ping_Failures})
        Total_Stats.update({'Ping_Success_rate': Ping_Success_rate})

        df_total_stats = pd.DataFrame(list(Total_Stats.items()), columns=['Statistic', 'Value'])
        # print(df_total_stats)
        df_sel['Session_No'] = df_sel['Session_ID'].cat.codes

        # endregion

        agg_str = lambda x: ','.join(filter(lambda s: pd.notna(s), x))

        agg_list = {'CombinedColumn': agg_str, 'Serving Cell RS SINR (dB)': 'mean', 'Serving Cell RSRP (dBm)': 'mean', 'Serving Cell RSRQ (dB)': 'mean', 'HTTP_URL': agg_str, 'HTTP_Download_Average_Throughput': 'mean', 'HTTP_Download_Service_Average_Throughput': 'mean', 'HTTP_Download_Session_Failure_Ratio': 'mean', 'HTTP_Download_Session_Success_Ratio': 'mean', 'HTTP_Outcome': agg_str, "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A": 'mean',
            "HTTP_Download_Data_Transfer_Success_Ratio_Method_A": 'mean', "HTTP_Download_Data_Transfer_Time_sec_Method_A": 'mean', "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "HTTP_Download_IP_Service_Setup_Time_sec_Method_A": 'mean', "HTTP_Download_Mean_Data_Rate_kbps_Method_A": 'mean', "HTTP_Download_Transfer_Start_Delay_Method_A": 'mean', "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B": 'mean',
            "HTTP_Download_Data_Transfer_Success_Ratio_Method_B": 'mean', "HTTP_Download_Data_Transfer_Time_sec_Method_B": 'mean', "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B": 'mean', "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "HTTP_Download_IP_Service_Setup_Time_sec_Method_B": 'mean', "HTTP_Download_Mean_Data_Rate_kbps_Method_B": 'mean', "HTTP_Upload_Average_Throughput": 'mean', "HTTP_Upload_Service_Average_Throughput": 'mean',
            "HTTP_Upload_Service_Transfer_Time": 'mean', "HTTP_Upload_Session_Failure_Ratio": 'mean', "HTTP_Upload_Session_Success_Ratio": 'mean', "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Time_sec_Method_A": 'mean', "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A": 'mean',
            "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A": 'mean', "HTTP_Upload_Mean_Data_Rate_kbps_Method_A": 'mean', "Streaming_URL": agg_str, "Streaming_Outcome_Type": agg_str, "Aborted_by_User": 'mean', "Streaming_Average_Session_Resolution": 'mean', "Streaming_Average_Throughput": 'mean', "Streaming_Completion_Rate": 'mean', "Streaming_Duration": 'mean', "Streaming_HD_Resolution": agg_str, "Streaming_HD_Resolution_Ratio": 'mean', "Streaming_Impairment_Free": agg_str,
            "Streaming_Impairment_Free_Video_Session_Ratio": 'mean', "Streaming_Maximum_Duration_Of_Video_Session_Interruptions": 'mean', "Streaming_Number_Of_Video_Session_Interruptions": 'mean', "Streaming_Player_Size_kB": 'mean', "Streaming_Reproduction_Cutoff_Ratio": 'mean', "Streaming_Reproduction_Start_Delay_sec": 'mean', "Streaming_Reproduction_Start_Failure_Ratio": 'mean', "Streaming_Service_Access_Time_ms": 'mean', "Streaming_Service_Access_Time_sec": 'mean',
            "Streaming_Session_Failure_Ratio": 'mean', "Streaming_Session_Qualified": agg_str, "Streaming_Session_Qualified_Ratio": 'mean', "Streaming_Session_Video_Interruption_Duration": 'mean', "Streaming_Session_Without_Interruption_Rate": 'mean', "Streaming_Setup_Success_Rate": 'mean', "Streaming_State_Prebuffering_to_Reproducing_Delay": 'mean', "Streaming_State_Request_to_Prebuffering_Delay": 'mean', "Streaming_State_Request_to_Reproducing_Delay": 'mean', "Streaming_Success_Rate": 'mean',
            "Streaming_Throughput_Filtered": 'mean', "Streaming_Total_Duration_Of_Video_Session_Interruptions": 'mean', "Streaming_Video_Buffer_Size_kB": 'mean', "Streaming_Video_IP_Service_Access_Time_ms": 'mean', "Streaming_Video_IP_Service_Access_Time_sec": 'mean', "Streaming_Video_Play_Start_Failure_Ratio": 'mean', "Streaming_Video_Play_Start_Time_sec": 'mean', "Streaming_Video_Session_Cutoff_Ratio": 'mean', "Streaming_Video_Session_Failure_Ratio": 'mean',
            "Streaming_Video_Session_Success_Ratio": 'mean', "Streaming_Video_Session_Time_sec": 'mean', "Streaming_Video_Size_kB": 'mean', "SessionTime": 'mean', "SessionTime_Upload": 'mean', "ThroughputCountOver1MBit": 'mean', "ThroughputCountOver3MBit": 'mean', "ThroughputPercentageOver3MBit": 'mean', "TCP_Handshake_Time_sec": 'mean', "Ping_Address": agg_str, "Ping_Count_Attempts": 'mean', "Ping_Count_Failed": 'mean', "Ping_Count_Success": 'mean', "Ping_Delay_ms_Avg": 'mean',
            "Ping_Delay_ms_Max": 'mean', "Ping_Delay_ms_Min": 'mean', "Ping_Packet_Loss_Rate": 'mean', "Ping_Packet_Success_Rate": 'mean', "Ping_Roundtrip_Time_ms": 'mean', "Ping_Roundtrip_Time_sec": 'mean', "Ping_Size": 'mean', "Data_Radio_Bearer": agg_str, "Fixed_Duration": 'mean', "IP_Interruption_Time_ms": 'mean', "Is_Multi_RAB": agg_str, "LTE_Serving_Cell_Count_Average": 'mean', "Radio_Access_Technology": agg_str, "Service_Bearer": agg_str,
            "FTP_Download_Average_Throughput_Not_Completed_Session": 'mean', "FTP_Download_Bearer": agg_str, "FTP_Download_Error_Cause": 'mean', "FTP_Download_Outcome": 'mean', "FTP_Download_Session_Failure_Ratio": 'mean', "FTP_Download_Session_Success_Ratio": 'mean', "FTP_Server_File": agg_str, "Seconds_Start_to_End_FTP_Download": 'mean', "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_A": 'mean', "FTP_Download_Data_Transfer_Success_Ratio_Method_A": 'mean',
            "FTP_Download_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "FTP_Download_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "FTP_Download_IP_Service_Setup_Time_sec_Method_A": 'mean', "FTP_Download_Transfer_Start_Delay_Method_A": 'mean', "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_B": 'mean', "FTP_Download_Data_Transfer_Success_Ratio_Method_B": 'mean', "FTP_Download_IP_Service_Access_Failure_Ratio_Method_B": 'mean',
            "FTP_Download_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "FTP_Download_IP_Service_Setup_Time_sec_Method_B": 'mean', "Seconds_Start_to_End_FTP_Upload": 'mean', "FTP_Upload_Bearer": agg_str, "FTP_Upload_Outcome": agg_str, "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "FTP_Upload_IP_Service_Setup_Time_sec_Method_A": 'mean', "FTP_Upload_Transfer_Start_Delay_Method_A": 'mean',
            "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_B": 'mean', "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "FTP_Upload_IP_Service_Setup_Time_sec_Method_B": 'mean', "Seconds_Start_to_End_UDP_Download": 'mean', "UDP_Download_Error_Cause": agg_str, "Seconds_Start_to_End_UDP_Upload": 'mean', "UDP_Upload_Error_Cause": agg_str, "DNS_Client": agg_str, "DNS_Domain_Name": agg_str, "DNS_First_In_Session": agg_str, "DNS_Host_Name_Resolution_Failure_Ratio": 'mean',
            "DNS_Host_Name_Resolution_Time_sec": 'mean', "DNS_Host_Name_Total_Resolution_Time_sec": 'mean', "DNS_Resolved_Address": agg_str, "DNS_Server_Address": agg_str}

        agg_list_2 = {'Latitude': ['min', 'max'], 'Longitude': ['min', 'max'], "Session_No": 'mean', "Test_type": agg_str, 'Distance': sum, 'CombinedColumn': agg_str, 'Serving Cell RS SINR (dB)': 'mean', 'Serving Cell RSRP (dBm)': 'mean', 'Serving Cell RSRQ (dB)': 'mean', 'HTTP_URL': agg_str, 'HTTP_Download_Average_Throughput': 'mean', 'HTTP_Download_Service_Average_Throughput': 'mean', 'HTTP_Download_Session_Failure_Ratio': 'mean', 'HTTP_Download_Session_Success_Ratio': 'mean',
            'HTTP_Outcome': agg_str, "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A": 'mean', "HTTP_Download_Data_Transfer_Success_Ratio_Method_A": 'mean', "HTTP_Download_Data_Transfer_Time_sec_Method_A": 'mean', "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "HTTP_Download_IP_Service_Setup_Time_sec_Method_A": 'mean', "HTTP_Download_Mean_Data_Rate_kbps_Method_A": 'mean',
            "HTTP_Download_Transfer_Start_Delay_Method_A": 'mean', "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B": 'mean', "HTTP_Download_Data_Transfer_Success_Ratio_Method_B": 'mean', "HTTP_Download_Data_Transfer_Time_sec_Method_B": 'mean', "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B": 'mean', "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "HTTP_Download_IP_Service_Setup_Time_sec_Method_B": 'mean', "HTTP_Download_Mean_Data_Rate_kbps_Method_B": 'mean',
            "HTTP_Upload_Average_Throughput": 'mean', "HTTP_Upload_Service_Average_Throughput": 'mean', "HTTP_Upload_Service_Transfer_Time": 'mean', "HTTP_Upload_Session_Failure_Ratio": 'mean', "HTTP_Upload_Session_Success_Ratio": 'mean', "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Time_sec_Method_A": 'mean', "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A": 'mean',
            "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A": 'mean', "HTTP_Upload_Mean_Data_Rate_kbps_Method_A": 'mean', "Streaming_URL": agg_str, "Streaming_Outcome_Type": agg_str, "Aborted_by_User": 'mean', "Streaming_Average_Session_Resolution": 'mean', "Streaming_Average_Throughput": 'mean', "Streaming_Completion_Rate": 'mean', "Streaming_Duration": 'mean', "Streaming_HD_Resolution": agg_str,
            "Streaming_HD_Resolution_Ratio": 'mean', "Streaming_Impairment_Free": agg_str, "Streaming_Impairment_Free_Video_Session_Ratio": 'mean', "Streaming_Maximum_Duration_Of_Video_Session_Interruptions": 'mean', "Streaming_Number_Of_Video_Session_Interruptions": 'mean', "Streaming_Player_Size_kB": 'mean', "Streaming_Reproduction_Cutoff_Ratio": 'mean', "Streaming_Reproduction_Start_Delay_sec": 'mean', "Streaming_Reproduction_Start_Failure_Ratio": 'mean',
            "Streaming_Service_Access_Time_ms": 'mean', "Streaming_Service_Access_Time_sec": 'mean', "Streaming_Session_Failure_Ratio": 'mean', "Streaming_Session_Qualified": agg_str, "Streaming_Session_Qualified_Ratio": 'mean', "Streaming_Session_Video_Interruption_Duration": 'mean', "Streaming_Session_Without_Interruption_Rate": 'mean', "Streaming_Setup_Success_Rate": 'mean', "Streaming_State_Prebuffering_to_Reproducing_Delay": 'mean', "Streaming_State_Request_to_Prebuffering_Delay": 'mean',
            "Streaming_State_Request_to_Reproducing_Delay": 'mean', "Streaming_Success_Rate": 'mean', "Streaming_Throughput_Filtered": 'mean', "Streaming_Total_Duration_Of_Video_Session_Interruptions": 'mean', "Streaming_Video_Buffer_Size_kB": 'mean', "Streaming_Video_IP_Service_Access_Time_ms": 'mean', "Streaming_Video_IP_Service_Access_Time_sec": 'mean', "Streaming_Video_Play_Start_Failure_Ratio": 'mean', "Streaming_Video_Play_Start_Time_sec": 'mean',
            "Streaming_Video_Session_Cutoff_Ratio": 'mean', "Streaming_Video_Session_Failure_Ratio": 'mean', "Streaming_Video_Session_Success_Ratio": 'mean', "Streaming_Video_Session_Time_sec": 'mean', "Streaming_Video_Size_kB": 'mean', "SessionTime": 'mean', "SessionTime_Upload": 'mean', "ThroughputCountOver1MBit": 'mean', "ThroughputCountOver3MBit": 'mean', "ThroughputPercentageOver3MBit": 'mean', "TCP_Handshake_Time_sec": 'mean', "Ping_Address": agg_str, "Ping_Count_Attempts": 'mean',
            "Ping_Count_Failed": 'mean', "Ping_Count_Success": 'mean', "Ping_Delay_ms_Avg": 'mean', "Ping_Delay_ms_Max": 'mean', "Ping_Delay_ms_Min": 'mean', "Ping_Packet_Loss_Rate": 'mean', "Ping_Packet_Success_Rate": 'mean', "Ping_Roundtrip_Time_ms": 'mean', "Ping_Roundtrip_Time_sec": 'mean', "Ping_Size": 'mean', "Data_Radio_Bearer": agg_str, "Fixed_Duration": 'mean', "IP_Interruption_Time_ms": 'mean', "Is_Multi_RAB": agg_str, "LTE_Serving_Cell_Count_Average": 'mean',
            "Radio_Access_Technology": agg_str, "Service_Bearer": agg_str, "FTP_Download_Average_Throughput_Not_Completed_Session": 'mean', "FTP_Download_Bearer": agg_str, "FTP_Download_Error_Cause": 'mean', "FTP_Download_Outcome": 'mean', "FTP_Download_Session_Failure_Ratio": 'mean', "FTP_Download_Session_Success_Ratio": 'mean', "FTP_Server_File": agg_str, "Seconds_Start_to_End_FTP_Download": 'mean', "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_A": 'mean',
            "FTP_Download_Data_Transfer_Success_Ratio_Method_A": 'mean', "FTP_Download_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "FTP_Download_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "FTP_Download_IP_Service_Setup_Time_sec_Method_A": 'mean', "FTP_Download_Transfer_Start_Delay_Method_A": 'mean', "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_B": 'mean', "FTP_Download_Data_Transfer_Success_Ratio_Method_B": 'mean', "FTP_Download_IP_Service_Access_Failure_Ratio_Method_B": 'mean',
            "FTP_Download_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "FTP_Download_IP_Service_Setup_Time_sec_Method_B": 'mean', "Seconds_Start_to_End_FTP_Upload": 'mean', "FTP_Upload_Bearer": agg_str, "FTP_Upload_Outcome": agg_str, "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_A": 'mean', "FTP_Upload_IP_Service_Setup_Time_sec_Method_A": 'mean', "FTP_Upload_Transfer_Start_Delay_Method_A": 'mean',
            "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_B": 'mean', "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_B": 'mean', "FTP_Upload_IP_Service_Setup_Time_sec_Method_B": 'mean', "Seconds_Start_to_End_UDP_Download": 'mean', "UDP_Download_Error_Cause": agg_str, "Seconds_Start_to_End_UDP_Upload": 'mean', "UDP_Upload_Error_Cause": agg_str, "DNS_Client": agg_str, "DNS_Domain_Name": agg_str, "DNS_First_In_Session": agg_str, "DNS_Host_Name_Resolution_Failure_Ratio": 'mean',
            "DNS_Host_Name_Resolution_Time_sec": 'mean', "DNS_Host_Name_Total_Resolution_Time_sec": 'mean', "DNS_Resolved_Address": agg_str, "DNS_Server_Address": agg_str}

        from haversine import haversine, Unit


        def calculate_distance(row):
            prev_coords = (row['prev_latitude'], row['prev_longitude'])
            current_coords = (row['Latitude'], row['Longitude'])
            return haversine(prev_coords, current_coords, unit=Unit.KILOMETERS)


        df_sel['prev_latitude'] = df_sel['Latitude'].shift(1)
        df_sel['prev_longitude'] = df_sel['Longitude'].shift(1)
        df_sel['Distance'] = df_sel.apply(calculate_distance, axis=1)
        df_sel['Distance'][0] = 0
        df_sel = df_sel.drop(['prev_latitude', 'prev_longitude'], axis=1)


        pivot_result = pd.pivot_table(df_sel, index=["Session_ID", "Test_ID", "Test_type"], values=list(agg_list.keys()), aggfunc=agg_list)
        df_pivot = pivot_result.reset_index()

        pivot_result_2 = pd.pivot_table(df_sel, index=["Test_ID"], values=list(agg_list_2.keys()), aggfunc=agg_list_2)

        pivot_result_2.columns = [' '.join(map(str, tpl)) for tpl in pivot_result_2.columns]
        pivot_result_2.columns = [s.replace(" <lambda>", '') for s in pivot_result_2.columns]
        pivot_result_2.columns = [s.replace(" mean", '') for s in pivot_result_2.columns]
        df_pivot = pivot_result_2.reset_index()

        def clear_duplicate_values_in_cell(cell_value):
            values = cell_value.split(',')
            unique_values = list(set(values))
            return ','.join(unique_values)


        def clear_duplicate_values_in_column(dataframe, column_name):
            df_copy = dataframe.copy()
            df_copy[column_name] = df_copy[column_name].apply(clear_duplicate_values_in_cell)
            return df_copy


        str_metrics = ["HTTP_URL", "HTTP_Outcome", "Streaming_URL", "Streaming_Outcome_Type", "Streaming_HD_Resolution", "Streaming_Impairment_Free", "Streaming_Session_Qualified", "Ping_Address", "Data_Radio_Bearer", "Is_Multi_RAB", "Radio_Access_Technology", "Service_Bearer", "FTP_Download_Bearer", "FTP_Server_File", "FTP_Upload_Bearer", "FTP_Upload_Outcome", "UDP_Download_Error_Cause", "UDP_Upload_Error_Cause", "DNS_Client", "DNS_Domain_Name", "DNS_First_In_Session", "DNS_Resolved_Address",
            "DNS_Server_Address"]

        df_result = clear_duplicate_values_in_column(df_pivot, "DNS_Client")
        df_result = clear_duplicate_values_in_column(df_result, "DNS_Domain_Name")
        df_result = clear_duplicate_values_in_column(df_result, "DNS_Resolved_Address")
        df_result = clear_duplicate_values_in_column(df_result, "DNS_Server_Address")
        df_result = clear_duplicate_values_in_column(df_result, "CombinedColumn")
        df_result = clear_duplicate_values_in_column(df_result, "Test_type")

        df_result = df_result.rename(columns={'CombinedColumn': 'Grouping_All_URL'})

        Grouping_All_Url = ["8.8.8.8", "Source Uri: https://www.youtube.com/watch?v=BQwC_DJSdfE", "URI: http://212.183.159.230/5MB.zip", "URI: http://d26kjwdsl72dzf.cloudfront.net/upload", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/Kepler_mobile/index.html", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/upload/", "URI: https://ash-speed.hetzner.com/1GB.bin", "URI: https://de.m.wikipedia.org/wiki/Europa", "URI: https://www.amazon.de/",
            "URI: https://www.chip.de/", "URI: https://www.gmx.net/", "URI: https://www.google.de/", "URI: https://www.instagram.com/"]
        Direction_Test = ["PING", "DL", "DL", "UL", "DL", "UL", "DL", "DL", "DL", "DL", "DL", "DL", "DL"]
        Test_Name = ["PING", "Streaming YT", "FDFS HTTP DL ST", "FDTT HTTP UL MT", "Kepler", "FDFS HTTP UL ST", "FDTT HTTP DL MT", "Wikipedia", "Amazon.de", "Chip.de", "Gmx.net", "Google.de", "Instagram.com"]
        Type_of_Test = ["PING", "Streaming YT", "FDFS DL", "FDTT UL ", "HTTP Browsing", "FDFS UL", "FDTT DL", "HTTP Browsing", "HTTP Browsing", "HTTP Browsing", "HTTP Browsing", "HTTP Browsing", "HTTP Browsing"]
        Thread_info = ["PING", "Live 4K", "ST", "MT", "MT", "ST", "MT", "MT", "MT", "MT", "MT", "MT", "MT"]

        data = {'Grouping_All_URL': Grouping_All_Url, 'Direction_Test': Direction_Test, 'Test_Name': Test_Name, 'Type_of_Test': Type_of_Test, 'Thread_info': Thread_info}
        df_gr = pd.DataFrame(data)

        merged_df = pd.merge(df_result, df_gr, on='Grouping_All_URL', how='left')

        columns_to_move = ['Grouping_All_URL', 'Direction_Test', 'Test_Name', 'Type_of_Test', 'Thread_info']
        new_positions = [3, 4, 5, 6, 7]  # 0-indexed positions

        # Move each column to its new position
        for column_name, new_position in zip(columns_to_move, new_positions):
            column_to_move = merged_df.pop(column_name)
            merged_df.insert(new_position, column_name, column_to_move)

        # Presenting results
        merged_df.to_csv('Agg_Results.csv', mode='w', header=True, index=False)
        # webbrowser.open("Pivot.csv")

        # region excel report creator

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side, Font

        Name_list = ["FDFS_DL_Attempts", "FDFS_DL_Success", "FDFS_DL_Failure", "FDFS_DL_Success_Ratio", "FDFS_UL_Attempts", "FDFS_UL_Success", "FDFS_UL_Failure", "FDFS_UL_Success_Ratio", "HTTP_Browsing_Attempts", "HTTP_Browsing_Access_Failure", "HTTP_Browsing_Access_Success", "HTTP_Browsing_Min_Transfer_Time_s", "HTTP_Browsing_10_PCTL_Transfer_Time_s", "HTTP_Browsing_Average_Transfer_Time_s", "HTTP_Browsing_Median_Transfer_Time_s", "HTTP_Browsing_90_PCTL_Transfer_Time_s",
            "HTTP_Browsing_Max_Transfer_Time_s", "Video_Stream_Attempts", "Video_Stream_Failures", "Video_Stream_Access_Time_sec", "Video_Stream_Success_Ratio_PC", "Video_Stream_Reproduction_Start_Delay_sec", "Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay", "Streaming_Reproduction_Cut-off_Ratio", "Streaming_Session_Without_Interruption_Rate_PC", "Streaming_Aggregated_Average_Session_Resolution_p", "Ping_Attempt", "Ping_Success", "Ping_Failures", "Ping_Success_rate"]
        column_headers = ["Campaign", "Operator", "Type_of_Test", "Test_Name", "Test_Info", "Direction", "Thread_Info", "URL", "Size", "Test_Result", "ErrorCode_Message", "Mean_Data_Rate_Mbit_s", "Transfer_Duration_s", "TestId", "Sessionid", "FileID", "Sequenz_ID_per_File", "File_Name", "PCAP_File_Name", "Channel", "Channel_Description", "G_Level_1", "G_Level_2", "G_Level_3", "G_Level_4", "G_Level_5", "Test_Start_Latitude", "Test_Start_Longitude", "Test_Distance_m", "Year", "Quarter", "Month",
            "Week", "Day", "Hour", "Test_Start_Time", "DNS_1st_Request", "DNS_1st_Response", "TCP_1st_SYN", "TCP_1st_SYN_ACK", "Data_1st_recieved", "Data_Last_Recieved", "Test_End_Time", "Test_Duration_s", "Transfer_Access_Duration_s", "Transfer_Transferred_Bytes", "FDTT_Sustainable_Transferred_Bytes", "FDTT_Sustainable_Transferred_Time_ms", "FDTT_Sustainable_MDR_Mbit_s", "http_Browser_Access_Duration_s", "http_Browser_Transferred_Bytes", "http_Browser_Number_of_Images",
            "http_Browser_1MB_Reached_Duration_ms", "Content_Transferred_Bytes", "IP_Service_Access_Result", "IP_Service_Access_Delay_ms", "DNS_Service_Access_Delay_ms", "TCP_RTT_Service_Access_Delay_ms", "Data_Download_Delay (SYNACK -> 1stData)", "APN_String", "Source_IP", "DNS_Server_IPs", "Destination_IP", "Threads", "TCP_Threads_Detailed", "DNS_Resolution_Attempts", "DNS_Resolution_Success", "DNS_Min_Resolution_Time_ms", "DNS_Avg_Resolution_Time_ms", "DNS_Max_Resolution_Time_ms",
            "DNS_Hosts_Resolved", "IP_Layer_Transferred_Bytes_DL", "IP_Layer_Transferred_Bytes_UL", "PCell_RAT", "PCell_RAT_Timeline", "TEC_Timeline", "TIME_GSM_s", "TIME_UMTS_s", "TIME_LTE_s", "TIME_LTE_CA_s", "TIME_NR5G_s", "TIME_No_Service_s", "IMEI", "IMSI", "MSISDN", "Device", "Firmware", "Measurement_System", "SW_Version", "Home_Operator", "Home_MCC", "Home_MNC", "MCC", "MNC", "CellID", "LAC", "BCCH", "ActiveSet", "SC1", "BSIC", "PCI", "LAC_CId_BCCH", "MinRSRP", "AvgRSRP", "MaxRSRP",
            "MinRSRQ", "AvgRSRQ", "MaxRSRQ", "MinSINR", "AvgSINR", "MaxSINR", "MinTxPwr4G", "AvgTxPwr4G", "MaxTxPwr4G", "CA_Configured", "CA_Configured_Timeline", "CA_Activated", "CA_Bands", "Test_Bandwith_Average_MHz", "PCell_Info_List", "PCell_Throughput_Average_Mbit_s", "SCell1_Info_List", "SCell1_Usage_Ratio", "SCell1_Throughput_Average_Mbit_s", "SCell2_Info_List", "SCell2_Usage_Ratio", "SCell2_Throughput_Average_Mbit_s", "CQI_LTE_Min", "CQI_LTE", "CQI_LTE_Max", "AvgDLMCS", "ShareDLQPSK",
            "ShareDL16QAM", "ShareDL64QAM", "ShareDL256QAM", "DL256QAM_larger_10", "PDSCHBytesTransfered", "MinNetPDSCHThroughput_Mbit_s", "AvgNetPDSCHThroughput_Mbit_s", "MaxNetPDSCHThroughput_Mbit_s", "ShareULBPSK", "ShareULQPSK", "ShareUL16QAM", "ShareUL64QAM", "ShareUL256QAM", "UL64QAM_larger_30", "PUSCHBytesTransfered", "MinNetPUSCHThroughput_Mbit_s", "AvgNetPUSCHThroughput_Mbit_s", "MaxNetPUSCHThroughput_Mbit_s", "NR_RSRP_MIN", "NR_RSRP_AVG", "NR_RSRP_MAX", "NR_RSRQ_MIN", "NR_RSRQ_AVG",
            "NR_RSRQ_MAX", "NR_SINR_MIN", "NR_SINR_AVG", "NR_SINR_MAX", "NR_TxPwr_MIN", "NR_TxPwr_AVG", "NR_TxPwr_MAX", "LTE_NR_Carriers", "LTE_NR_Band_Combination", "NrDl_ARFCN", "NrDl_PointA", "NrDl_PCI", "NrDl_Band", "NrDl_Total_Time", "NrB3_Total_Time", "NrDl_Scheduled_PDSCH_Throughput_Avg", "NrDl_Net_PDSCH_Throughput_Avg", "NrDl_Total_Bandwidth", "NrDl_Serving_Beam_Index", "NrDl_Numerologies_Number", "Numerology1_SCS", "Numerology1_BW", "NrDl_Initial_BWP_Bandwidth", "NrDl_Initial_BWP_SCS",
            "NrDl_BWP1_Bandwidth", "NrDl_BWP1_SCS", "NrDl_Active_BWP", "NrDl_CQI_Avg", "NrDl_MCS_Avg", "NrDl_QPSK_Ratio", "NrDl_16QAM_Ratio", "NrDl_64QAM_Ratio", "NrDl_256QAM_Ratio", "NrDl_MinTBSize", "NrDl_AvgTBSize", "NrDl_MaxTBSize", "NrDl_MinTBRate", "NrDl_AvgTBRate", "NrDl_MaxTBRate", "NrDl_MinRBs", "NrDl_AvgRBs", "NrDl_MaxRBs", "NrDl_Rank", "NrDl_Max_MIMO_Layers", "NrUl_ARFCN", "NrUl_Total_Bandwidth", "NrUl_Scheduled_PUSCH_Throughput_Avg_Mbit_s", "NrUl_Net_PUSCH_Throughput_Avg_Mbit_s",
            "NrUl_Initial_BWP_Bandwidth", "NrUl_Initial_BWP_SCS", "NrUl_BWP1_Bandwidth", "NrUl_BWP1_SCS", "NrUl_Active_BWP", "NrUl_MCS_Avg", "NrUl_QPSK_Ratio", "NrUl_16QAM_Ratio", "NrUl_64QAM_Ratio", "NrUl_256QAM_Ratio", "NrUl_MinTBSize", "NrUl_AvgTBSize", "NrUl_MaxTBSize", "NrUl_MinTBRate", "NrUl_AvgTBRate", "NrUl_MaxTBRate", "NrUl_MinRBs", "NrUl_AvgRBs", "NrUl_MaxRBs", "NrUl_Rank", "NrUl_Max_MIMO_Layers", "HandoversInfo", "Region", "Vendor", "JOIN_ID", "VDF_CELL_NAME", "BatteryLevel",
            "BatteryTemp", "CPUTemp", "CPULoad"]

        data1 = {'Name': Name_list}
        data2 = {'Name': column_headers}

        df1 = df_total_stats  # pd.DataFrame(data1)
        df2 = pd.DataFrame(columns=column_headers)

        # region creating vodafone table] = merged_df[""]
        # df2["Campaign"] = merged_df[""]
        df2["Operator"] = "Vodafone"
        df2["Type_of_Test"] = merged_df["Type_of_Test"]
        # df2["Test_Name"] = merged_df[""]
        # df2["Test_Info"] = merged_df[""]
        df2["Direction"] = merged_df["Direction_Test"]
        df2["Thread_Info"] = merged_df["Thread_info"]
        df2["URL"] = merged_df["Grouping_All_URL"]
        # df2["Size"] = merged_df[""]
        # df2["Test_Result"] = merged_df[""]
        # df2["ErrorCode_Message"] = merged_df[""]
        # df2["Mean_Data_Rate_Mbit_s"] = merged_df[""]
        # df2["Transfer_Duration_s"] = merged_df[""]
        df2["TestId"] = merged_df["Test_ID"]
        df2["Sessionid"] = merged_df["Session_No"]
        # df2["FileID"] = merged_df[""]
        df2["Sequenz_ID_per_File"] = merged_df["Test_ID"]
        # df2["File_Name"] = merged_df[""]
        # df2["PCAP_File_Name"] = merged_df[""]
        # df2["Channel"] = merged_df[""]
        # df2["Channel_Description"] = merged_df[""]
        # df2["G_Level_1"] = merged_df[""]
        # df2["G_Level_2"] = merged_df[""]
        # df2["G_Level_3"] = merged_df[""]
        # df2["G_Level_4"] = merged_df[""]
        # df2["G_Level_5"] = merged_df[""]
        df2["Test_Start_Latitude"] = merged_df["Latitude min"]
        df2["Test_Start_Longitude"] = merged_df["Longitude min"]
        df2["Test_Distance_m"] = merged_df["Distance sum"]
        df2["Year"] = 2024
        # df2["Quarter"] = merged_df[""]
        # df2["Month"] = merged_df[""]
        # df2["Week"] = merged_df[""]
        # df2["Day"] = merged_df[""]
        # df2["Hour"] = merged_df[""]
        # df2["Test_Start_Time"] = merged_df[""]
        # df2["DNS_1st_Request"] = merged_df[""]
        # df2["DNS_1st_Response"] = merged_df[""]
        # df2["TCP_1st_SYN"] = merged_df[""]
        # df2["TCP_1st_SYN_ACK"] = merged_df[""]
        # df2["Data_1st_recieved"] = merged_df[""]
        # df2["Data_Last_Recieved"] = merged_df[""]
        # df2["Test_End_Time"] = merged_df[""]
        # df2["Test_Duration_s"] = merged_df[""]
        # df2["Transfer_Access_Duration_s"] = merged_df[""]
        # df2["Transfer_Transferred_Bytes"] = merged_df[""]
        # df2["FDTT_Sustainable_Transferred_Bytes"] = merged_df[""]
        # df2["FDTT_Sustainable_Transferred_Time_ms"] = merged_df[""]
        # df2["FDTT_Sustainable_MDR_Mbit_s"] = merged_df[""]
        # df2["http_Browser_Access_Duration_s"] = merged_df[""]
        # df2["http_Browser_Transferred_Bytes"] = merged_df[""]
        # df2["http_Browser_Number_of_Images"] = merged_df[""]
        # df2["http_Browser_1MB_Reached_Duration_ms"] = merged_df[""]
        # df2["Content_Transferred_Bytes"] = merged_df[""]
        # df2["IP_Service_Access_Result"] = merged_df[""]
        # df2["IP_Service_Access_Delay_ms"] = merged_df[""]
        # df2["DNS_Service_Access_Delay_ms"] = merged_df[""]
        # df2["TCP_RTT_Service_Access_Delay_ms"] = merged_df[""]
        # df2["Data_Download_Delay (SYNACK -> 1stData)"] = merged_df[""]
        # df2["APN_String"] = merged_df[""]
        # df2["Source_IP"] = merged_df[""]
        # df2["DNS_Server_IPs"] = merged_df[""]
        # df2["Destination_IP"] = merged_df[""]
        # df2["Threads"] = merged_df[""]
        # df2["TCP_Threads_Detailed"] = merged_df[""]
        # df2["DNS_Resolution_Attempts"] = merged_df[""]
        # df2["DNS_Resolution_Success"] = merged_df[""]
        # df2["DNS_Min_Resolution_Time_ms"] = merged_df[""]
        # df2["DNS_Avg_Resolution_Time_ms"] = merged_df[""]
        # df2["DNS_Max_Resolution_Time_ms"] = merged_df[""]
        # df2["DNS_Hosts_Resolved"] = merged_df[""]
        # df2["IP_Layer_Transferred_Bytes_DL"] = merged_df[""]
        # df2["IP_Layer_Transferred_Bytes_UL"] = merged_df[""]
        # df2["PCell_RAT"] = merged_df[""]
        # df2["PCell_RAT_Timeline"] = merged_df[""]
        # df2["TEC_Timeline"] = merged_df[""]
        # df2["TIME_GSM_s"] = merged_df[""]
        # df2["TIME_UMTS_s"] = merged_df[""]
        # df2["TIME_LTE_s"] = merged_df[""]
        # df2["TIME_LTE_CA_s"] = merged_df[""]
        # df2["TIME_NR5G_s"] = merged_df[""]
        # df2["TIME_No_Service_s"] = merged_df[""]
        # df2["IMEI"] = merged_df[""]
        # df2["IMSI"] = merged_df[""]
        # df2["MSISDN"] = merged_df[""]
        # df2["Device"] = merged_df[""]
        # df2["Firmware"] = merged_df[""]
        # df2["Measurement_System"] = merged_df[""]
        # df2["SW_Version"] = merged_df[""]
        # df2["Home_Operator"] = merged_df[""]
        # df2["Home_MCC"] = merged_df[""]
        # df2["Home_MNC"] = merged_df[""]
        # df2["MCC"] = merged_df[""]
        # df2["MNC"] = merged_df[""]
        # df2["CellID"] = merged_df[""]
        # df2["LAC"] = merged_df[""]
        # df2["BCCH"] = merged_df[""]
        # df2["ActiveSet"] = merged_df[""]
        # df2["SC1"] = merged_df[""]
        # df2["BSIC"] = merged_df[""]
        # df2["PCI"] = merged_df[""]
        # df2["LAC_CId_BCCH"] = merged_df[""]
        # df2["MinRSRP"] = merged_df[""]
        df2["AvgRSRP"] = merged_df["Serving Cell RSRP (dBm)"]
        # df2["MaxRSRP"] = merged_df[""]
        # df2["MinRSRQ"] = merged_df[""]
        df2["AvgRSRQ"] = merged_df["Serving Cell RSRQ (dB)"]
        # df2["MaxRSRQ"] = merged_df[""]
        # df2["MinSINR"] = merged_df[""]
        df2["AvgSINR"] = merged_df["Serving Cell RS SINR (dB)"]
        # df2["MaxSINR"] = merged_df[""]
        # df2["MinTxPwr4G"] = merged_df[""]
        # df2["AvgTxPwr4G"] = merged_df[""]
        # df2["MaxTxPwr4G"] = merged_df[""]
        # df2["CA_Configured"] = merged_df[""]
        # df2["CA_Configured_Timeline"] = merged_df[""]
        # df2["CA_Activated"] = merged_df[""]
        # df2["CA_Bands"] = merged_df[""]
        # df2["Test_Bandwith_Average_MHz"] = merged_df[""]
        # df2["PCell_Info_List"] = merged_df[""]
        # df2["PCell_Throughput_Average_Mbit_s"] = merged_df[""]
        # df2["SCell1_Info_List"] = merged_df[""]
        # df2["SCell1_Usage_Ratio"] = merged_df[""]
        # df2["SCell1_Throughput_Average_Mbit_s"] = merged_df[""]
        # df2["SCell2_Info_List"] = merged_df[""]
        # df2["SCell2_Usage_Ratio"] = merged_df[""]
        # df2["SCell2_Throughput_Average_Mbit_s"] = merged_df[""]
        # df2["CQI_LTE_Min"] = merged_df[""]
        # df2["CQI_LTE"] = merged_df[""]
        # df2["CQI_LTE_Max"] = merged_df[""]
        # df2["AvgDLMCS"] = merged_df[""]
        # df2["ShareDLQPSK"] = merged_df[""]
        # df2["ShareDL16QAM"] = merged_df[""]
        # df2["ShareDL64QAM"] = merged_df[""]
        # df2["ShareDL256QAM"] = merged_df[""]
        # df2["DL256QAM_larger_10"] = merged_df[""]
        # df2["PDSCHBytesTransfered"] = merged_df[""]
        # df2["MinNetPDSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["AvgNetPDSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["MaxNetPDSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["ShareULBPSK"] = merged_df[""]
        # df2["ShareULQPSK"] = merged_df[""]
        # df2["ShareUL16QAM"] = merged_df[""]
        # df2["ShareUL64QAM"] = merged_df[""]
        # df2["ShareUL256QAM"] = merged_df[""]
        # df2["UL64QAM_larger_30"] = merged_df[""]
        # df2["PUSCHBytesTransfered"] = merged_df[""]
        # df2["MinNetPUSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["AvgNetPUSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["MaxNetPUSCHThroughput_Mbit_s"] = merged_df[""]
        # df2["NR_RSRP_MIN"] = merged_df[""]
        # df2["NR_RSRP_AVG"] = merged_df[""]
        # df2["NR_RSRP_MAX"] = merged_df[""]
        # df2["NR_RSRQ_MIN"] = merged_df[""]
        # df2["NR_RSRQ_AVG"] = merged_df[""]
        # df2["NR_RSRQ_MAX"] = merged_df[""]
        # df2["NR_SINR_MIN"] = merged_df[""]
        # df2["NR_SINR_AVG"] = merged_df[""]
        # df2["NR_SINR_MAX"] = merged_df[""]
        # df2["NR_TxPwr_MIN"] = merged_df[""]
        # df2["NR_TxPwr_AVG"] = merged_df[""]
        # df2["NR_TxPwr_MAX"] = merged_df[""]
        # df2["LTE_NR_Carriers"] = merged_df[""]
        # df2["LTE_NR_Band_Combination"] = merged_df[""]
        # df2["NrDl_ARFCN"] = merged_df[""]
        # df2["NrDl_PointA"] = merged_df[""]
        # df2["NrDl_PCI"] = merged_df[""]
        # df2["NrDl_Band"] = merged_df[""]
        # df2["NrDl_Total_Time"] = merged_df[""]
        # df2["NrB3_Total_Time"] = merged_df[""]
        # df2["NrDl_Scheduled_PDSCH_Throughput_Avg"] = merged_df[""]
        # df2["NrDl_Net_PDSCH_Throughput_Avg"] = merged_df[""]
        # df2["NrDl_Total_Bandwidth"] = merged_df[""]
        # df2["NrDl_Serving_Beam_Index"] = merged_df[""]
        # df2["NrDl_Numerologies_Number"] = merged_df[""]
        # df2["Numerology1_SCS"] = merged_df[""]
        # df2["Numerology1_BW"] = merged_df[""]
        # df2["NrDl_Initial_BWP_Bandwidth"] = merged_df[""]
        # df2["NrDl_Initial_BWP_SCS"] = merged_df[""]
        # df2["NrDl_BWP1_Bandwidth"] = merged_df[""]
        # df2["NrDl_BWP1_SCS"] = merged_df[""]
        # df2["NrDl_Active_BWP"] = merged_df[""]
        # df2["NrDl_CQI_Avg"] = merged_df[""]
        # df2["NrDl_MCS_Avg"] = merged_df[""]
        # df2["NrDl_QPSK_Ratio"] = merged_df[""]
        # df2["NrDl_16QAM_Ratio"] = merged_df[""]
        # df2["NrDl_64QAM_Ratio"] = merged_df[""]
        # df2["NrDl_256QAM_Ratio"] = merged_df[""]
        # df2["NrDl_MinTBSize"] = merged_df[""]
        # df2["NrDl_AvgTBSize"] = merged_df[""]
        # df2["NrDl_MaxTBSize"] = merged_df[""]
        # df2["NrDl_MinTBRate"] = merged_df[""]
        # df2["NrDl_AvgTBRate"] = merged_df[""]
        # df2["NrDl_MaxTBRate"] = merged_df[""]
        # df2["NrDl_MinRBs"] = merged_df[""]
        # df2["NrDl_AvgRBs"] = merged_df[""]
        # df2["NrDl_MaxRBs"] = merged_df[""]
        # df2["NrDl_Rank"] = merged_df[""]
        # df2["NrDl_Max_MIMO_Layers"] = merged_df[""]
        # df2["NrUl_ARFCN"] = merged_df[""]
        # df2["NrUl_Total_Bandwidth"] = merged_df[""]
        # df2["NrUl_Scheduled_PUSCH_Throughput_Avg_Mbit_s"] = merged_df[""]
        # df2["NrUl_Net_PUSCH_Throughput_Avg_Mbit_s"] = merged_df[""]
        # df2["NrUl_Initial_BWP_Bandwidth"] = merged_df[""]
        # df2["NrUl_Initial_BWP_SCS"] = merged_df[""]
        # df2["NrUl_BWP1_Bandwidth"] = merged_df[""]
        # df2["NrUl_BWP1_SCS"] = merged_df[""]
        # df2["NrUl_Active_BWP"] = merged_df[""]
        # df2["NrUl_MCS_Avg"] = merged_df[""]
        # df2["NrUl_QPSK_Ratio"] = merged_df[""]
        # df2["NrUl_16QAM_Ratio"] = merged_df[""]
        # df2["NrUl_64QAM_Ratio"] = merged_df[""]
        # df2["NrUl_256QAM_Ratio"] = merged_df[""]
        # df2["NrUl_MinTBSize"] = merged_df[""]
        # df2["NrUl_AvgTBSize"] = merged_df[""]
        # df2["NrUl_MaxTBSize"] = merged_df[""]
        # df2["NrUl_MinTBRate"] = merged_df[""]
        # df2["NrUl_AvgTBRate"] = merged_df[""]
        # df2["NrUl_MaxTBRate"] = merged_df[""]
        # df2["NrUl_MinRBs"] = merged_df[""]
        # df2["NrUl_AvgRBs"] = merged_df[""]
        # df2["NrUl_MaxRBs"] = merged_df[""]
        # df2["NrUl_Rank"] = merged_df[""]
        # df2["NrUl_Max_MIMO_Layers"] = merged_df[""]
        # df2["HandoversInfo"] = merged_df[""]
        # df2["Region"] = merged_df[""]
        # df2["Vendor"] = merged_df[""]
        # df2["JOIN_ID"] = merged_df[""]
        # df2["VDF_CELL_NAME"] = merged_df[""]
        # df2["BatteryLevel"] = merged_df[""]
        # df2["BatteryTemp"] = merged_df[""]
        # df2["CPUTemp"] = merged_df[""]
        # df2["CPULoad"] = merged_df[""]

        # endregion

        # Specify the Excel file name
        excel_file = 'output_file.xlsx'

        # Create an Excel writer
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            # Write each DataFrame to a different sheet
            df1.to_excel(writer, sheet_name='Master', index=False)
            df2.to_excel(writer, sheet_name='Vodafone', index=False)

        wb = load_workbook(excel_file)

        sheet = wb["Master"]
        # Adjust the width of the first column
        sheet.column_dimensions['A'].width = 60  # Set the width to 15

        sheet = wb["Vodafone"]
        sheet.column_dimensions['A'].width = 30  # Set the width to 15

        # Save the changes to the Excel file
        wb.save(excel_file)
        #print(f'Excel file "{excel_file}" created and styled successfully.')
        #import webbrowser
        #webbrowser.open(excel_file)


        # endregion

        def convert_df(df):
            # IMPORTANT: Cache the conversion to prevent computation on every rerun
            return df.to_csv(index=False).encode('utf-8')

        csv_1 = convert_df(df_total_stats)
        csv = convert_df(merged_df)
        xlsx_1 = 'output_file.xlsx'
        st.download_button(label="Download Result File_statistics", data=csv_1, file_name="Agg_Stats.csv", mime='text/csv',key=4)
        st.download_button(label="Download Result File_pivot", data=csv, file_name="Agg_Pivot.csv", mime='text/csv',key=5)
        # st.download_button(label="Download Result File_excel_report", data= xlsx_1, file_name='output_file.xlsx', mime='text/csv',key=6)
        st.download_button(label="Download Result File_excel_report", key=6, data=open('output_file.xlsx', 'rb').read(), file_name='output_file.xlsx')
        st.dataframe(df_total_stats)
        st.dataframe(merged_df)


    st.divider()  # 👈 Draws a horizontal rule

    st.header('Drive Test Statistical Report Module - Voice')

    uploaded_file = st.file_uploader("Choose input zip file to process", type="zip", accept_multiple_files=True,key = 1)

    if st.button('Process uploaded files', key = 2):

        with zipfile.ZipFile(uploaded_file[0], 'r') as zip_ref:
            csv_file_name = zip_ref.namelist()[0]
            zip_ref.extract(csv_file_name)


        # Script start

        # path = "C:/08_2024_DT/HiDrive-CDR/Master Data/O2_CDR_DATA_RATINGEN_20240207/"
        # csv_file_name = path + "Metric Group 1.csv"
        df = pd.read_csv(csv_file_name, low_memory=False)

if selected == "CDR Reports Historical Comparison":
    st.title('Comparison Module for Drive Tests')
    st.header('Drive Test Statistical Report Module')
    st.write("This module is a demo for DR Reports Historical Comparison ")

if selected == "View Historical Reports":
    st.title('Report Module for Historical Drive Tests')
    st.header('Historical Reports viewing Module')
    st.write("This module is a demo for CDR Reports Historical Vieving ")

if selected == "View on Map":
    st.title('Map Module for Historical Drive Tests')
    st.header('Historical Reports Map viewing Module')
    st.write("This module is a demo for CDR Reports Map Vieving ")



    uploaded_file_2 = st.file_uploader("Choose input zip file to process", type="zip", accept_multiple_files=True)

    if st.button('Process uploaded files'):
        import pandas as pd
        import pydeck as pdk
        import zipfile

        with zipfile.ZipFile(uploaded_file_2[0], 'r') as zip_ref:
            csv_file_name = zip_ref.namelist()[0]
            zip_ref.extract(csv_file_name)


        df = pd.read_csv(csv_file_name, low_memory=False).head(50000)

        sel_columns = ["Time", "Date", "Latitude", "Longitude", "Grouping", "Grouping_with_Direction", "Grouping_with_Direction_HTTP", "Grouping_with_Direction_Ping", "Operator", "Task_Type", "Technology", "HTTP_URL", "Streaming_URL", "Ping_Address", "Serving Cell RS SINR (dB)", "Serving Cell RSRP (dBm)", "Serving Cell RSRQ (dB)", "HTTP_Download_Average_Throughput", "HTTP_Download_Service_Average_Throughput", "HTTP_Download_Session_Failure_Ratio", "HTTP_Download_Session_Success_Ratio",
            "HTTP_Outcome", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Download_Data_Transfer_Success_Ratio_Method_A", "HTTP_Download_Data_Transfer_Time_sec_Method_A", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Time_sec_Method_A", "HTTP_Download_Mean_Data_Rate_kbps_Method_A", "HTTP_Download_Transfer_Start_Delay_Method_A", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B",
            "HTTP_Download_Data_Transfer_Success_Ratio_Method_B", "HTTP_Download_Data_Transfer_Time_sec_Method_B", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Time_sec_Method_B", "HTTP_Download_Mean_Data_Rate_kbps_Method_B", "HTTP_Upload_Average_Throughput", "HTTP_Upload_Service_Average_Throughput", "HTTP_Upload_Service_Transfer_Time", "HTTP_Upload_Session_Failure_Ratio",
            "HTTP_Upload_Session_Success_Ratio", "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A"]

        # for i in sel_columns:
        # print(i)

        df_sel = df[sel_columns]

        # st.write(df_sel.columns)

        lon_cen_1 = df_sel["Longitude"].mean()
        lat_cen_1 = df_sel["Latitude"].mean()

        view_state = pdk.ViewState(latitude=lat_cen_1, longitude=lon_cen_1, zoom=10)

        st.pydeck_chart(pdk.Deck(map_style=None, initial_view_state=view_state,
            layers=[
            pdk.Layer('ScatterplotLayer',
                data=df_sel,
                get_position='[Longitude, Latitude]',
                get_color='[200, 30, 0, 160]',
                pickable=True,
                tooltip=True,
                radiusScale=2,
                radiusMinPixels=2,
                radiusMaxPixels=2) # get_radius=50,
            ],
            # tooltip={"Name: {name}"}
            tooltip={"text": "HTTP_Download_Service_Average_Throughput: {HTTP_Download_Service_Average_Throughput}"}
        ))





