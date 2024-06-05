def data_preprocessing(df):
    import pandas as pd
    import numpy as np

    columns_to_combine = ["HTTP_URL", "Streaming_URL", "Ping_Address"]

    sel_columns_old = ["Time", "Date", "Latitude", "Longitude", "Grouping", "Grouping_with_Direction", "Grouping_with_Direction_HTTP", "Grouping_with_Direction_Ping", "Operator", "Task_Type", "Technology", "HTTP_URL", "Streaming_URL", "Ping_Address", "Serving Cell RS SINR (dB)", "Serving Cell RSRP (dBm)", "Serving Cell RSRQ (dB)", "HTTP_Download_Average_Throughput", "HTTP_Download_Service_Average_Throughput", "HTTP_Download_Session_Failure_Ratio", "HTTP_Download_Session_Success_Ratio",
        "HTTP_Outcome", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Download_Data_Transfer_Success_Ratio_Method_A", "HTTP_Download_Data_Transfer_Time_sec_Method_A", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Time_sec_Method_A", "HTTP_Download_Mean_Data_Rate_kbps_Method_A", "HTTP_Download_Transfer_Start_Delay_Method_A", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B",
        "HTTP_Download_Data_Transfer_Success_Ratio_Method_B", "HTTP_Download_Data_Transfer_Time_sec_Method_B", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Time_sec_Method_B", "HTTP_Download_Mean_Data_Rate_kbps_Method_B", "HTTP_Upload_Average_Throughput", "HTTP_Upload_Service_Average_Throughput", "HTTP_Upload_Service_Transfer_Time", "HTTP_Upload_Session_Failure_Ratio",
        "HTTP_Upload_Session_Success_Ratio", "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Time_sec_Method_A", "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A", "HTTP_Upload_Mean_Data_Rate_kbps_Method_A", "HTTP Transfer Time (second)", "Streaming_Outcome_Type", "Aborted_by_User",
        "Streaming_Average_Session_Resolution", "Streaming_Average_Throughput", "Streaming_Completion_Rate", "Streaming_Duration", "Streaming_HD_Resolution", "Streaming_HD_Resolution_Ratio", "Streaming_Impairment_Free", "Streaming_Impairment_Free_Video_Session_Ratio", "Streaming_Maximum_Duration_Of_Video_Session_Interruptions", "Streaming_Number_Of_Video_Session_Interruptions", "Streaming_Player_Size_kB", "Streaming_Reproduction_Cutoff_Ratio", "Streaming_Reproduction_Start_Delay_sec",
        "Streaming_Reproduction_Start_Failure_Ratio", "Streaming_Service_Access_Time_ms", "Streaming_Service_Access_Time_sec", "Streaming_Session_Failure_Ratio", "Streaming_Session_Qualified", "Streaming_Session_Qualified_Ratio", "Streaming_Session_Video_Interruption_Duration", "Streaming_Session_Without_Interruption_Rate", "Streaming_Setup_Success_Rate", "Streaming_State_Prebuffering_to_Reproducing_Delay", "Streaming_State_Request_to_Prebuffering_Delay",
        "Streaming_State_Request_to_Reproducing_Delay", "Streaming_Success_Rate", "Streaming_Throughput_Filtered", "Streaming_Total_Duration_Of_Video_Session_Interruptions", "Streaming_Video_Buffer_Size_kB", "Streaming_Video_IP_Service_Access_Time_ms", "Streaming_Video_IP_Service_Access_Time_sec", "Streaming_Video_Play_Start_Failure_Ratio", "Streaming_Video_Play_Start_Time_sec", "Streaming_Video_Session_Cutoff_Ratio", "Streaming_Video_Session_Failure_Ratio",
        "Streaming_Video_Session_Success_Ratio", "Streaming_Video_Session_Time_sec", "Streaming_Video_Size_kB", # "Streaming_Impairment_Free",
        # "Streaming_Reproduction_Cutoff_Ratio",
        "SessionTime", "SessionTime_Upload", "ThroughputCountOver1MBit", "ThroughputCountOver3MBit", "ThroughputPercentageOver3MBit", "TCP_Handshake_Time_sec", "Ping_Count_Attempts", "Ping_Count_Failed", "Ping_Count_Success", "Ping_Delay_ms_Avg", "Ping_Delay_ms_Max", "Ping_Delay_ms_Min", "Ping_Packet_Loss_Rate", "Ping_Packet_Success_Rate", "Ping_Roundtrip_Time_ms", "Ping_Roundtrip_Time_sec", "Ping_Size", "Data_Radio_Bearer", "Fixed_Duration", "IP_Interruption_Time_ms", "Is_Multi_RAB",
        "LTE_Serving_Cell_Count_Average", "Radio_Access_Technology", "Service_Bearer", "FTP_Download_Average_Throughput_Not_Completed_Session", "FTP_Download_Bearer", "FTP_Download_Error_Cause", "FTP_Download_Outcome", "FTP_Download_Session_Failure_Ratio", "FTP_Download_Session_Success_Ratio", "FTP_Server_File", "Seconds_Start_to_End_FTP_Download", "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_A", "FTP_Download_Data_Transfer_Success_Ratio_Method_A",
        "FTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "FTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "FTP_Download_IP_Service_Setup_Time_sec_Method_A", "FTP_Download_Transfer_Start_Delay_Method_A", "FTP_Download_Data_Transfer_Cutoff_Ratio_Method_B", "FTP_Download_Data_Transfer_Success_Ratio_Method_B", "FTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "FTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "FTP_Download_IP_Service_Setup_Time_sec_Method_B",
        "Seconds_Start_to_End_FTP_Upload", "FTP_Upload_Bearer", "FTP_Upload_Outcome", "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_A", "FTP_Upload_IP_Service_Setup_Time_sec_Method_A", "FTP_Upload_Transfer_Start_Delay_Method_A", "FTP_Upload_IP_Service_Access_Failure_Ratio_Method_B", "FTP_Upload_IP_Service_Setup_Success_Ratio_Method_B", "FTP_Upload_IP_Service_Setup_Time_sec_Method_B", "Seconds_Start_to_End_UDP_Download",
        "UDP_Download_Error_Cause", "Seconds_Start_to_End_UDP_Upload", "UDP_Upload_Error_Cause", "DNS_Client", "DNS_Domain_Name", "DNS_First_In_Session", "DNS_Host_Name_Resolution_Failure_Ratio", "DNS_Host_Name_Resolution_Time_sec", "DNS_Host_Name_Total_Resolution_Time_sec", "DNS_Resolved_Address", "DNS_Server_Address"]
    sel_columns_old_2 = [
# "Campaign",
# "Operator",
"Date Time",
"Latitude",
"Longitude",
"Operator",
"File Name",
"Grouping",
"Size",
"Test_Result",
"ErrorCode_Message",
"Mean_Data_Rate_Mbit_s",
"Transfer_Duration_s",
"TestId",
"Sessionid",
"FileID",
"HTTP_URL",
"HTTP_Outcome",
"HTTP_Download_Average_Throughput",
"HTTP_Download_Service_Average_Throughput",
"HTTP_Download_Session_Failure_Ratio",
"HTTP_Download_Session_Success_Ratio",
"HTTP Data Transfer Cutoff Ratio (%)",
"HTTP Download Throughput (kbps)",
"HTTP_Download_Data_Transfer_Failure_Ratio_Method_A",
"HTTP_Download_Data_Transfer_Success_Ratio_Method_A",
"HTTP_Download_Data_Transfer_Time_sec_Method_A",
"HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A",
"HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A",
"HTTP_Download_IP_Service_Setup_Time_sec_Method_A",
"HTTP_Download_Mean_Data_Rate_kbps_Method_A",
"HTTP_Download_Transfer_Start_Delay_Method_A",
"HTTP_Download_Data_Transfer_Failure_Ratio_Method_B",
"HTTP_Download_Data_Transfer_Success_Ratio_Method_B",
"HTTP_Download_Data_Transfer_Time_sec_Method_B",
"HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B",
"HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B",
"HTTP_Download_IP_Service_Setup_Time_sec_Method_B",
"HTTP_Download_Mean_Data_Rate_kbps_Method_B",
"HTTP_Upload_Average_Throughput",
"HTTP_Upload_Service_Average_Throughput",
"HTTP_Upload_Service_Transfer_Time",
"HTTP_Upload_Session_Failure_Ratio",
"HTTP_Upload_Session_Success_Ratio",
"HTTP Upload Throughput Current (kbps)",
"HTTP Upload Throughput Mean (kbps)",
"HTTP Upload Transfer Time (second)",
"HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A",
"HTTP_Upload_Data_Transfer_Success_Ratio_Method_A",
"HTTP_Upload_Data_Transfer_Time_sec_Method_A",
"HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A",
"HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A",
"HTTP_Upload_IP_Service_Setup_Time_sec_Method_A",
"HTTP_Upload_Mean_Data_Rate_kbps_Method_A",
"HTTP Transfer Time (second)",
"HTTP IP Service Access Failure Ratio (%)",
"HTTP Get Session Failure Ratio (%)",
"HTTP Post Session Failure Ratio (%)",
"http_Browser_Access_Duration_s",
"http_Browser_Transferred_Bytes",
"http_Browser_Number_of_Images",
"http_Browser_1MB_Reached_Duration_ms",
"Streaming_URL",
"Streaming_Outcome_Type",
"Aborted_by_User",
"Streaming_Average_Session_Resolution",
"Streaming_Average_Throughput",
"Streaming_Completion_Rate",
"Streaming_Duration",
"Streaming_HD_Resolution",
"Streaming_HD_Resolution_Ratio",
"Streaming_Impairment_Free",
"Streaming_Impairment_Free_Video_Session_Ratio",
"Streaming_Maximum_Duration_Of_Video_Session_Interruptions",
"Streaming_Number_Of_Video_Session_Interruptions",
"Streaming_Player_Size_kB",
"Streaming_Reproduction_Cutoff_Ratio",
"Streaming_Reproduction_Start_Delay_sec",
"Streaming_Reproduction_Start_Failure_Ratio",
"Streaming_Service_Access_Time_ms",
"Streaming_Service_Access_Time_sec",
"Streaming_Session_Failure_Ratio",
"Streaming_Session_Qualified",
"Streaming_Session_Qualified_Ratio",
"Streaming_Session_Video_Interruption_Duration",
"Streaming_Session_Without_Interruption_Rate",
"Streaming_Setup_Success_Rate",
"Streaming_State_Prebuffering_to_Reproducing_Delay",
"Streaming_State_Request_to_Prebuffering_Delay",
"Streaming_State_Request_to_Reproducing_Delay",
"Streaming_Success_Rate",
"Streaming_Throughput_Filtered",
"Streaming_Total_Duration_Of_Video_Session_Interruptions",
"Streaming_Video_Buffer_Size_kB",
"Streaming_Video_Play_Start_Failure_Ratio",
"Streaming_Video_Play_Start_Time_sec",
"Streaming_Video_Session_Cutoff_Ratio",
"Streaming_Video_Session_Failure_Ratio",
"Streaming_Video_Session_Success_Ratio",
"Streaming_Video_Session_Time_sec",
"Streaming_Video_Size_kB",
"Streaming Packet Loss",
"Streaming Throughput (kbps)",
"Streaming Aggregated Average Session Resolution",
"Streaming Audio Bit Rate",
"Streaming Frame Height",
"Streaming Frame Width",
"Streaming Player Download Data Transfer Failure Ratio",
"Streaming Player Service IP Access Failure Ratio",
"Streaming Player Session Failure Ratio",
"Streaming Session Cutoff Ratio",
"Streaming Success Ratio",
"Streaming Video Best Resolution",
"Streaming Video Bit Rate",
"Streaming Video Download Data Transfer Failure Ratio",
"Streaming Video Interruption Count",
"Streaming Video Interruption Duration",
"Streaming Video Interruption Duration Ratio",
"Streaming Video Service IP Access Failure Ratio",
"VideoStream_VQ_MOS",
"VideoStream_Video_Resolution",
"VideoStream_Video_Resolution_Timeline",
"VideoStream_FrameRate",
"VideoStream_Number_of_Freezings",
"VideoStream_Freezing_Time_Sum_s",
"VideoStream_Freezing_Time_Max_s",
"VideoStream_Freezing_Proportion",
"VideoStream_StreamedTime_s",
"VideoStream_Video_Codec",
"VideoStream_Dynamic_Range",
"VideoStream_BitDepth",
"VideoStream_BitRate(Mbps)",
"VideoStream_FHD_Ratio",
"VideoStream_UHD_Ratio",
"VideoStream_Black_Frames_Ratio",
"VideoStream_Jerkiness_Ratio",
"VideoStream_Number_of_Skips",
"VideoStream_Skipped_Time_Sum_s",
"Irritating_Video_Playout",
"VideoStream_Time_To_Start_Buffering_ms",
"VideoStream_PreBuffering_Time_ms",
"VideoStream_Time_to_First_Picture_s",
"VideoStream_Video_IP_Service_Access_Time_s",
"VideoStream_Video_Reproduction_Start_Delay_s",
"VideoStream_Video_Play_Start_Time_s",
"VideoStream_Video_Transfer_Time_s",
"VideoStream_Video_Playout_Duration_Time_s",
"VideoStream_Video_Playout_Cutoff_Time_s",
"VideoStream_Video_Downloaded_Size_kbit",
"VideoStream_Video_Stream_Duration_s",
"VideoStream_Video_IP_Service_Access_Timestamp",
"VideoStream_Video_Transfer_Start_Timestamp",
"VideoStream_Video_Reproduction_Start_Timestamp",
"VideoStream_Video_Play_Start_Timestamp",
"VideoStream_Video_Session_End_Timestamp",
"Content_Transferred_Bytes",
"Transfer_Access_Duration_s",
"Transfer_Transferred_Bytes",
"DNS_Client",
"DNS_Domain_Name",
"DNS_First_In_Session",
"DNS_Host_Name_Resolution_Failure_Ratio",
"DNS_Host_Name_Resolution_Time_sec",
"DNS_Host_Name_Total_Resolution_Time_sec",
"DNS_Resolved_Address",
"DNS_Server_Address",
"DNS_Resolution_Attempts",
"DNS_Resolution_Success",
"DNS_Min_Resolution_Time_ms",
"DNS_Avg_Resolution_Time_ms",
"DNS_Max_Resolution_Time_ms",
"DNS_Hosts_Resolved",
"DNS_Resolution_Success_Ratio (%)",
"DNS_1st_Request",
"DNS_1st_Response",
"DNS_Service_Access_Delay_ms",
"DNS_Server_IPs",
"Destination_IP",
"Source_IP",
"IP_Service_Access_Result",
"IP_Service_Access_Delay_ms",
"IP_Layer_Transferred_Bytes_DL",
"IP_Layer_Transferred_Bytes_UL",
"IP_Interruption_Time_ms",
"TCP_1st_SYN",
"TCP_1st_SYN_ACK",
"TCP_RTT_Service_Access_Delay_ms",
"TCP_Threads",
"TCP_Threads_Detailed",
"TCP_Handshake_Time_sec",
"TLS Handshake Failure Ratio (%)",
"Data_Download_Delay (SYNACK -> 1stData)",
"APN_String",
"Threads",
"ThroughputCountOver1MBit",
"ThroughputCountOver3MBit",
"ThroughputPercentageOver3MBit",
"IE Browse Download Time (ms)",
"IE Browse Page Size (Bytes)",
"Ping_Address",
"Ping_Count_Attempts",
"Ping_Count_Failed",
"Ping_Count_Success",
"Ping_Delay_ms_Avg",
"Ping_Delay_ms_Max",
"Ping_Delay_ms_Min",
"Ping_Packet_Loss_Rate",
"Ping_Packet_Success_Rate",
"Ping_Roundtrip_Time_ms",
"Ping_Roundtrip_Time_sec",
"Ping_Size",
"Application Layer Throughput Downlink (kbps)",
"Application Layer Throughput Downlink (Mbps)",
"Application Layer Throughput Uplink (kbps)",
"Application Layer Throughput Uplink (Mbps)",
"SessionTime",
"SessionTime_Upload",
"Multi_RAT_Phy_Throughput_DL_kbps",
"Multi_RAT_Phy_Throughput_DL_Mbps",
"Multi_RAT_Phy_Throughput_UL_kbps",
"Multi RAT Connectivity Mode",
"Is_Multi_RAB",
"RAT",
"PCell_RAT",
"PCell_RAT_Timeline",
"TEC_Timeline",
"Mode",
"Data_Radio_Bearer",
"Service_Bearer",
"Fixed_Duration",
"LTE_Serving_Cell_Count_Average",
"Serving Cell Channel RSSI (dBm)",
"Serving Cell DL EARFCN",
"Serving Cell eMTC Cell Indication",
"Serving Cell Frequency Band",
"Serving Cell Identity",
"Serving Cell Serving Cell EN-DC Support",
"Serving Beam 1 Band",
"Serving Beam 1 Bandwidth DL",
"Serving Beam 1 Cell Identity",
"Serving Beam 1 Cell Type",
"Serving Beam 1 GSCN",
"Serving Beam 1 NRARFCN DL",
"Serving Beam 1 SSB Index",
"MAC DL Throughput (kbps)",
"MAC DL Throughput Carrier 1 (kbps)",
"MAC DL Throughput Carrier 2 (kbps)",
"MAC DL Throughput Carrier 3 (kbps)",
"MAC DL Throughput Carrier 4 (kbps)",
"MAC UL Throughput (kbps)",
"MAC UL Throughput Carrier 1 (kbps)",
"MAC UL Throughput Carrier 2 (kbps)",
"PDSCH Modulation C1 TB0",
"PDSCH CW0 BLER Carrier 1 (%)",
"PDSCH Phy Throughput (kbps)",
"PDSCH RI SB",
"PDSCH BLER (%): BLER SB",
"PDSCH Modulation CW0: MOD SB",
"PDSCH Phy Throughput Total (kbps)",
"PDSCH Phy Throughput Total (Mbps)",
"NR Phy Throughput Multi-RAT DL (kbps)",
"NR Phy Throughput Multi-RAT DL (Mbps)",
"PDSCHBytesTransferred",
"NetPDSCHThroughput_kbps",
"MinDLRB",
"AvgDLRB",
"MaxDLRB",
"AvgDLMCSIndex",
"ShareDLQPSK",
"ShareDL16QAM",
"ShareDL64QAM",
"ShareDL256QAM",
"DL256QAM_larger_10",
"PUSCH Phy Throughput (kbps)",
"PUSCH BLER Carrier 1 (%)",
"PUSCHBytesTransferred",
"NetPUSCHThroughput_Kbps",
"ShareULBPSK",
"ShareULQPSK",
"ShareUL16QAM",
"ShareUL64QAM",
"UL64QAM_larger_30",
"NR Phy Throughput Multi-RAT UL (kbps)",
"NR Phy Throughput Multi-RAT UL (Mbps)",
"NR_RSRP (dBm)",
"NR_RSRQ (dB)",
"NR SINR (dB)",
"NR_TxPwr",
"LTE_NR_Carriers",
"LTE_NR_Band_Combination",
"NrDl_ARFCN",
"NrDl_PCI",
"NrDl_Band",
"NrDl_Total_Time",
"NrDl_Scheduled_PDSCH_Throughput_Avg",
"NrDl_Net_PDSCH_Throughput_Avg",
"NrDl_Total_Bandwidth",
"NrDl_Serving_Beam_Index",
"NrDl_Numerologies_Number",
"Numerology1_SCS",
"Numerology1_BW",
"NrDl_Initial_BWP_Bandwidth",
"NrDl_Initial_BWP_Subcarrier_Spacing",
"NrDl_BWP1_Bandwidth",
"NrDl_BWP1_SCS",
"NrDl_Active_BWP",
"NrDl_CQI_Avg",
"NrDl_MCS_Avg",
"NrDl_QPSK_Ratio",
"NrDl_16QAM_Ratio",
"NrDl_64QAM_Ratio",
"NrDl_256QAM_Ratio",
"NrDl_TBSize",
"NrDl_TBRate",
"NrDl_RBs",
"NrDl_Rank",
"NrDl_Max_MIMO_Layers",
"NrUl_ARFCN",
"NrUl_Total_Bandwidth",
"NrUl_Scheduled_PUSCH_Throughput_Avg_Mbit_s",
"NrUl_Net_PUSCH_Throughput_Avg_Mbit_s",
"NrUl_Initial_BWP_Bandwidth",
"NrUl_Initial_BWP_SCS",
"NrUl_BWP1_Bandwidth",
"NrUl_BWP1_SCS",
"NrUl_Active_BWP",
"NrUl_MCS_Avg",
"NrUl_QPSK_Ratio",
"NrUl_16QAM_Ratio",
"NrUl_64QAM_Ratio",
"NrUl_256QAM_Ratio",
"NrUl_TBSize",
"NrUl_TBRate",
"NrUl_RBs",
"NrUl_Rank",
"NrUl_Max_MIMO_Layers",
"CA_Configured",
"CA_Configured_Timeline",
"CA_Activated",
"CA_Bands",
"Test_Bandwith_Average_MHz",
"PCell_Info_List",
"PCell_Throughput_Average_Mbit_s",
"CQI_LTE",
"LTE_RSRP (dBm)",
"LTE_RSRQ (dB)",
"LTE_SINR (dB)",
"TxPwr4G",
"MinTxPwr4G",
"AvgTxPwr4G",
"MaxTxPwr4G",
"TxPwr3G",
"MinTxPwr3G",
"AvgTxPwr3G",
"MaxTxPwr3G",
"CQI_HSDPA_Min",
"CQI_HSDPA_",
"CQI_HSDPA_Max",
"BLER3G",
"EcIo",
"RSCP",
"RxQual",
"RxLev",
"TIME_GSM_s_Idle",
"TIME_WCDMA_s",
"TIME_LTE_s",
"TIME_LTE_CA_s",
"TIME_NR5G_s",
"TIME_No_Service_s",
"IMEI",
"IMSI",
"MSISDN",
"Device",
"Firmware",
"Measurement_System",
"SW_Version",
"Home_Operator",
"Home_MCC",
"Home_MNC",
"MCC",
"MNC",
"CellID",
"LAC",
"BCCH",
"ActiveSet",
"SC1",
"BSIC",
"PCI",
"LAC_CId_BCCH",
"Channel",
"Channel_Description",
"Test_Start_Latitude",
"Test_Start_Longitude",
"Test_Distance_m",
"Year",
"Quarter",
"Month",
"Week",
"Day",
"Hour",
"Test_Start_Time"]
    sel_columns_old_3 = ["Date Time", "Latitude", "Longitude", "Grouping", "Grouping_with_Direction", "Grouping_with_Direction_HTTP", "Grouping_with_Direction_Ping", "Technology_Detail", "Operator", "Logfile_Name", "IMEI", "IMSI", "MCC", "MNC", "Network Operator", "HTTP_URL", "HTTP_Outcome", "HTTP_Download_Average_Throughput", "HTTP_Download_Service_Average_Throughput", "HTTP_Download_Session_Failure_Ratio", "HTTP_Download_Session_Success_Ratio", "HTTP Data Transfer Cutoff Ratio (%)",
        "HTTP Download Throughput (kbps)", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Download_Data_Transfer_Success_Ratio_Method_A", "HTTP_Download_Data_Transfer_Time_sec_Method_A", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Time_sec_Method_A", "HTTP_Download_Mean_Data_Rate_kbps_Method_A", "HTTP_Download_Transfer_Start_Delay_Method_A",
        "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B", "HTTP_Download_Data_Transfer_Success_Ratio_Method_B", "HTTP_Download_Data_Transfer_Time_sec_Method_B", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Time_sec_Method_B", "HTTP_Download_Mean_Data_Rate_kbps_Method_B", "HTTP_Upload_Average_Throughput", "HTTP_Upload_Service_Average_Throughput", "HTTP_Upload_Service_Transfer_Time",
        "HTTP_Upload_Session_Failure_Ratio", "HTTP_Upload_Session_Success_Ratio", "HTTP Upload Throughput Current (kbps)", "HTTP Upload Throughput Mean (kbps)", "HTTP Upload Transfer Time (second)", "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Time_sec_Method_A", "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A",
        "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A", "HTTP_Upload_Mean_Data_Rate_kbps_Method_A", "HTTP Transfer Time (second)", "HTTP IP Service Access Failure Ratio (%)", "HTTP Get Session Failure Ratio (%)", "HTTP Post Session Failure Ratio (%)", "Seconds_Start_to_End_HTTP", "Seconds_Start_to_End_HTTP_Post", "HTTPS_Connection", "HTTP_Technical_Browsing_Time_sec", "HTTP_Upload_Bearer", "HTTP_Download_Bearer", "HTTP_Post_Outcome", "HTTP_Post_Error_Cause", "HTTP_User_Browsing_Time_sec",
        "Streaming_URL", "Streaming_Outcome_Type", "Aborted_by_User", "Streaming_Average_Session_Resolution", "Streaming_Average_Throughput", "Streaming_Completion_Rate", "Streaming_Duration", "Streaming_HD_Resolution", "Streaming_HD_Resolution_Ratio", "Streaming_Impairment_Free", "Streaming_Impairment_Free_Video_Session_Ratio", "Streaming_Maximum_Duration_Of_Video_Session_Interruptions", "Streaming_Number_Of_Video_Session_Interruptions", "Streaming_Player_Size_kB",
        "Streaming_Reproduction_Cutoff_Ratio", "Streaming_Reproduction_Start_Delay_sec", "Streaming_Reproduction_Start_Failure_Ratio", "Streaming_Service_Access_Time_ms", "Streaming_Service_Access_Time_sec", "Streaming_Session_Failure_Ratio", "Streaming_Session_Qualified", "Streaming_Session_Qualified_Ratio", "Streaming_Session_Video_Interruption_Duration", "Streaming_Session_Without_Interruption_Rate", "Streaming_Setup_Success_Rate", "Streaming_State_Prebuffering_to_Reproducing_Delay",
        "Streaming_State_Request_to_Prebuffering_Delay", "Streaming_State_Request_to_Reproducing_Delay", "Streaming_Success_Rate", "Streaming_Throughput_Filtered", "Streaming_Total_Duration_Of_Video_Session_Interruptions", "Streaming_Video_Buffer_Size_kB", "Streaming_Video_Play_Start_Failure_Ratio", "Streaming_Video_Play_Start_Time_sec", "Streaming_Video_Session_Cutoff_Ratio", "Streaming_Video_Session_Failure_Ratio", "Streaming_Video_Session_Success_Ratio", "Streaming_Video_Session_Time_sec",
        "Streaming_Video_Size_kB", "Streaming Packet Loss", "Streaming Throughput (kbps)", "Streaming Aggregated Average Session Resolution", "Streaming Audio Bit Rate", "Streaming Frame Height", "Streaming Frame Width", "Streaming Player Download Data Transfer Failure Ratio", "Streaming Player Service IP Access Failure Ratio", "Streaming Player Session Failure Ratio", "Streaming Session Cutoff Ratio", "Streaming Success Ratio", "Streaming Video Best Resolution", "Streaming Video Bit Rate",
        "Streaming Video Download Data Transfer Failure Ratio", "Streaming Video Interruption Count", "Streaming Video Interruption Duration", "Streaming Video Interruption Duration Ratio", "Streaming Video Service IP Access Failure Ratio", "Streaming_Video_IP_Service_Access_Time_sec", "Streaming Video Playout Duration (s)", "Streaming PEVQ-S Instantaneous MOS", "DNS_Client", "DNS_Domain_Name", "DNS_First_In_Session", "DNS_Host_Name_Resolution_Failure_Ratio", "DNS_Host_Name_Resolution_Time_sec",
        "DNS_Host_Name_Total_Resolution_Time_sec", "DNS_Resolved_Address", "DNS_Server_Address", "Resolution Success Ratio (%)", "First DNS Request", "IP_Interruption_Time_ms", "TCP_Handshake_Time_sec", "TLS Handshake Failure Ratio (%)", "ThroughputCountOver1MBit", "ThroughputCountOver3MBit", "ThroughputPercentageOver3MBit", "IE Browse Download Time (ms)", "IE Browse Page Size (Bytes)", "Ping_Address", "Ping_Count_Attempts", "Ping_Count_Failed", "Ping_Count_Success", "Ping_Delay_ms_Avg",
        "Ping_Delay_ms_Max", "Ping_Delay_ms_Min", "Ping_Packet_Loss_Rate", "Ping_Packet_Success_Rate", "Ping_Roundtrip_Time_ms", "Ping_Roundtrip_Time_sec", "Ping_Size", "Service_Bearer", "Fixed_Duration", "LTE_Serving_Cell_Count_Average", "Serving Cell Channel RSSI (dBm)", "Serving Cell DL EARFCN", "Serving Cell eMTC Cell Indication", "Serving Cell Frequency Band", "LTE Serving Cell Identity", "Serving Cell Serving Cell EN-DC Support", "Serving Cell RSRP (dBm)", "Serving Cell RSRQ (dB)",
        "Serving Cell RS SINR (dB)", "MAC DL Throughput (kbps)", "MAC DL Throughput Carrier 1 (kbps)", "MAC DL Throughput Carrier 2 (kbps)", "MAC DL Throughput Carrier 3 (kbps)", "MAC DL Throughput Carrier 4 (kbps)", "MAC UL Throughput (kbps)", "MAC UL Throughput Carrier 1 (kbps)", "MAC UL Throughput Carrier 2 (kbps)", "PDSCH Modulation C1 TB0", "PDSCH CW0 BLER Carrier 1 (%)", "PDSCH Phy Throughput (kbps)", "PDSCH Phy Throughput Total (kbps)", "PUSCH Phy Throughput (kbps)",
        "PUSCH BLER Carrier 1 (%)", "PDSCH Phy Throughput Carrier 1 (kbps)", "PDSCH Phy Throughput Carrier 2 (kbps)", "PDSCH Phy Throughput Carrier 3 (kbps)", "PDSCH Phy Throughput Carrier 4 (kbps)", "Serving Cell DL Pathloss Carrier 1 (dB)", "UE TX Power - PUSCH (dBm) Carrier 1", "UE TX Power - PUCCH (dBm) Carrier 1", "NR Phy Throughput Multi-RAT UL (kbps)", "NR Phy Throughput Multi-RAT UL (Mbps)", "NR Phy Throughput Multi-RAT DL (kbps)", "NR Phy Throughput Multi-RAT DL (Mbps)",
        "Serving Beam 1 NRARFCN DL", "Serving Beam 1 Cell Identity", "Serving Beam 1 Band", "Serving Beam 1 Bandwidth DL", "Serving Beam 1 SSB Index", "Serving Beam 1 Bandwidth UL", "Serving Beam 1 Cell Type", "Serving Beam 1 GSCN", "Serving Beam 1 SSB RSRP (dBm)", "Serving Beam 1 SSB RSRQ (dB)", "Serving Beam 1 SSB SNR (dB)", "DL Pathloss", "PUSCH Tx Power (dBm)", "PUCCH Tx Power (dBm)", "PDSCH BLER (%)", "PUSCH BLER (%)", "CQI", "PDSCH RI", "Application Layer Throughput Downlink (kbps)",
        "Application Layer Throughput Uplink (kbps)", "PUSCH Phy Throughput Total (kbps)", "NR MAC DL Throughput Total (kbps)", "NR MAC UL Throughput Total (kbps)", "NR RLC DL Throughput Total (kbps)", "NR RLC UL Throughput Total (kbps)", "MR-DC DRB PDCP DL Throughput Total (kbps)", "MR-DC DRB PDCP UL Throughput Total (kbps)", "PDSCH Phy Throughput Serving Beam 1 (kbps)", "PDSCH Phy Throughput Serving Beam 2 (kbps)", "PDSCH Phy Throughput Serving Beam 3 (kbps)",
        "PDSCH Phy Throughput Serving Beam 4 (kbps)", "BWP Bandwidth DL (Mhz)", "BWP Center NR-ARFCN DL", "BWP ID DL", "BWP Initial Bandwidth DL (MHz)", "BWP Initial Start Point NR-ARFCN DL", "BWP Start Point NR-ARFCN DL", "BWP Subcarrier Spacing DL", "BWPs Configured Count DL", "BWP Bandwidth UL (Mhz)", "BWP Center NR-ARFCN UL", "BWP ID UL", "BWP Initial Bandwidth UL (MHz)", "BWP Initial Start Point NR-ARFCN UL", "BWP Start Point NR-ARFCN UL", "BWP Subcarrier Spacing UL",
        "BWPs Configured Count UL", "Multi_RAT_Phy_Throughput_DL_kbps", "Multi_RAT_Phy_Throughput_DL_Mbps", "Multi_RAT_Phy_Throughput_UL_kbps", "Multi RAT Connectivity Mode", "Is_Multi_RAB", "RadioAccessTechnologyState"
    ]
    sel_columns = ["Date Time", "Latitude", "Longitude", "Grouping", "Grouping_with_Direction", "Grouping_with_Direction_HTTP", "Grouping_with_Direction_Ping", "Technology_Detail", "Operator", "Logfile_Name", "IMEI", "IMSI", "MCC", "MNC", "Network Operator", "SIM Operator", "HTTP_URL", "HTTP_Outcome", "HTTP_Download_Average_Throughput", "HTTP_Download_Service_Average_Throughput", "HTTP_Download_Session_Failure_Ratio", "HTTP_Download_Session_Success_Ratio", "HTTP Data Transfer Cutoff Ratio (%)",
        "HTTP Download Throughput (kbps)", "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Download_Data_Transfer_Success_Ratio_Method_A", "HTTP_Download_Data_Transfer_Time_sec_Method_A", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A", "HTTP_Download_IP_Service_Setup_Time_sec_Method_A", "HTTP_Download_Mean_Data_Rate_kbps_Method_A", "HTTP_Download_Transfer_Start_Delay_Method_A",
        "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B", "HTTP_Download_Data_Transfer_Success_Ratio_Method_B", "HTTP_Download_Data_Transfer_Time_sec_Method_B", "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B", "HTTP_Download_IP_Service_Setup_Time_sec_Method_B", "HTTP_Download_Mean_Data_Rate_kbps_Method_B", "HTTP_Upload_Average_Throughput", "HTTP_Upload_Service_Average_Throughput", "HTTP_Upload_Service_Transfer_Time",
        "HTTP_Upload_Session_Failure_Ratio", "HTTP_Upload_Session_Success_Ratio", "HTTP Upload Throughput Current (kbps)", "HTTP Upload Throughput Mean (kbps)", "HTTP Upload Transfer Time (second)", "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A", "HTTP_Upload_Data_Transfer_Time_sec_Method_A", "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A", "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A",
        "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A", "HTTP_Upload_Mean_Data_Rate_kbps_Method_A", "HTTP Transfer Time (second)", "HTTP IP Service Access Failure Ratio (%)", "HTTP Get Session Failure Ratio (%)", "HTTP Post Session Failure Ratio (%)", "Seconds_Start_to_End_HTTP", "Seconds_Start_to_End_HTTP_Post", "HTTPS_Connection", "HTTP_Technical_Browsing_Time_sec", "HTTP_Upload_Bearer", "HTTP_Download_Bearer", "HTTP_Post_Outcome", "HTTP_Post_Error_Cause", "HTTP_User_Browsing_Time_sec",
        "Streaming_URL", "Streaming_Outcome_Type", "Aborted_by_User", "Streaming_Average_Session_Resolution",
        "Streaming_Average_Throughput", "Streaming_Completion_Rate", "Streaming_Duration", "Streaming_HD_Resolution",
        "Streaming_HD_Resolution_Ratio", "Streaming_Impairment_Free", "Streaming_Impairment_Free_Video_Session_Ratio", "Streaming_Maximum_Duration_Of_Video_Session_Interruptions", "Streaming_Number_Of_Video_Session_Interruptions", "Streaming_Player_Size_kB",
        "Streaming_Reproduction_Cutoff_Ratio", "Streaming_Reproduction_Start_Delay_sec", "Streaming_Reproduction_Start_Failure_Ratio", "Streaming_Service_Access_Time_ms", "Streaming_Service_Access_Time_sec", "Streaming_Session_Failure_Ratio", "Streaming_Session_Qualified", "Streaming_Session_Qualified_Ratio", "Streaming_Session_Video_Interruption_Duration", "Streaming_Session_Without_Interruption_Rate", "Streaming_Setup_Success_Rate", "Streaming_State_Prebuffering_to_Reproducing_Delay",
        "Streaming_State_Request_to_Prebuffering_Delay", "Streaming_State_Request_to_Reproducing_Delay", "Streaming_Success_Rate", "Streaming_Throughput_Filtered", "Streaming_Total_Duration_Of_Video_Session_Interruptions", "Streaming_Video_Buffer_Size_kB", "Streaming_Video_Play_Start_Failure_Ratio", "Streaming_Video_Play_Start_Time_sec", "Streaming_Video_Session_Cutoff_Ratio", "Streaming_Video_Session_Failure_Ratio", "Streaming_Video_Session_Success_Ratio", "Streaming_Video_Session_Time_sec",
        "Streaming_Video_Size_kB", "Streaming Packet Loss", "Streaming Throughput (kbps)", "Streaming Aggregated Average Session Resolution", "Streaming Audio Bit Rate", "Streaming Frame Height", "Streaming Frame Width", "Streaming Player Download Data Transfer Failure Ratio", "Streaming Player Service IP Access Failure Ratio", "Streaming Player Session Failure Ratio", "Streaming Session Cutoff Ratio", "Streaming Success Ratio", "Streaming Video Best Resolution", "Streaming Video Bit Rate",
        "Streaming Video Download Data Transfer Failure Ratio", "Streaming Video Interruption Count", "Streaming Video Interruption Duration", "Streaming Video Interruption Duration Ratio", "Streaming Video Service IP Access Failure Ratio", "Streaming_Video_IP_Service_Access_Time_sec", "Streaming Video Playout Duration (s)", "Streaming PEVQ-S Instantaneous MOS", "DNS_Client", "DNS_Domain_Name", "DNS_First_In_Session", "DNS_Host_Name_Resolution_Failure_Ratio", "DNS_Host_Name_Resolution_Time_sec",
        "DNS_Host_Name_Total_Resolution_Time_sec", "DNS_Resolved_Address", "DNS_Server_Address", "Resolution Success Ratio (%)",
        "First DNS Request", "IP_Interruption_Time_ms", "TCP_Handshake_Time_sec", "TLS Handshake Failure Ratio (%)",
        "ThroughputCountOver1MBit", "ThroughputCountOver3MBit", "ThroughputPercentageOver3MBit", "IE Browse Download Time (ms)",
        "IE Browse Page Size (Bytes)",
        "Ping_Address", "Ping_Count_Attempts", "Ping_Count_Failed", "Ping_Count_Success",
        "Ping_Delay_ms_Avg",
        "Ping_Delay_ms_Max", "Ping_Delay_ms_Min", "Ping_Packet_Loss_Rate", "Ping_Packet_Success_Rate",
        "Ping_Roundtrip_Time_ms", "Ping_Roundtrip_Time_sec", "Ping_Size", "Service_Bearer", "Fixed_Duration",
        "LTE_Serving_Cell_Count_Average", "Serving Cell Channel RSSI (dBm)", "Serving Cell DL EARFCN", "Serving Cell eMTC Cell Indication", "Serving Cell Frequency Band", "LTE Serving Cell Identity", "Serving Cell Serving Cell EN-DC Support", "Serving Cell RSRP (dBm)", "Serving Cell RSRQ (dB)",
        "Serving Cell RS SINR (dB)", "MAC DL Throughput (kbps)", "MAC DL Throughput Carrier 1 (kbps)", "MAC DL Throughput Carrier 2 (kbps)", "MAC DL Throughput Carrier 3 (kbps)", "MAC DL Throughput Carrier 4 (kbps)", "MAC UL Throughput (kbps)", "MAC UL Throughput Carrier 1 (kbps)", "MAC UL Throughput Carrier 2 (kbps)", "PDSCH Modulation C1 TB0", "PDSCH CW0 BLER Carrier 1 (%)", "PDSCH Phy Throughput (kbps)", "PDSCH Phy Throughput Total (kbps)", "PUSCH Phy Throughput (kbps)",
        "PUSCH BLER Carrier 1 (%)", "PDSCH Phy Throughput Carrier 1 (kbps)", "PDSCH Phy Throughput Carrier 2 (kbps)", "PDSCH Phy Throughput Carrier 3 (kbps)", "PDSCH Phy Throughput Carrier 4 (kbps)", "Serving Cell DL Pathloss Carrier 1 (dB)", "UE TX Power - PUSCH (dBm) Carrier 1", "UE TX Power - PUCCH (dBm) Carrier 1", "NR Phy Throughput Multi-RAT UL (kbps)", "NR Phy Throughput Multi-RAT UL (Mbps)", "NR Phy Throughput Multi-RAT DL (kbps)", "NR Phy Throughput Multi-RAT DL (Mbps)",
        "Serving Beam 1 NRARFCN DL", "Serving Beam 1 Cell Identity", "Serving Beam 1 Band", "Serving Beam 1 Bandwidth DL", "Serving Beam 1 SSB Index", "Serving Beam 1 Bandwidth UL", "Serving Beam 1 Cell Type", "Serving Beam 1 GSCN", "Serving Beam 1 SSB RSRP (dBm)", "Serving Beam 1 SSB RSRQ (dB)", "Serving Beam 1 SSB SNR (dB)", "DL Pathloss", "PUSCH Tx Power (dBm)", "PUCCH Tx Power (dBm)", "PDSCH BLER (%)", "PUSCH BLER (%)", "CQI", "PDSCH RI", "Application Layer Throughput Downlink (kbps)",
        "Application Layer Throughput Uplink (kbps)", "PUSCH Phy Throughput Total (kbps)", "NR MAC DL Throughput Total (kbps)", "NR MAC UL Throughput Total (kbps)", "NR RLC DL Throughput Total (kbps)", "NR RLC UL Throughput Total (kbps)", "MR-DC DRB PDCP DL Throughput Total (kbps)", "MR-DC DRB PDCP UL Throughput Total (kbps)", "PDSCH Phy Throughput Serving Beam 1 (kbps)", "PDSCH Phy Throughput Serving Beam 2 (kbps)", "PDSCH Phy Throughput Serving Beam 3 (kbps)",
        "PDSCH Phy Throughput Serving Beam 4 (kbps)", "BWP Bandwidth DL (Mhz)", "BWP Center NR-ARFCN DL", "BWP ID DL", "BWP Initial Bandwidth DL (MHz)", "BWP Initial Start Point NR-ARFCN DL", "BWP Start Point NR-ARFCN DL", "BWP Subcarrier Spacing DL", "BWPs Configured Count DL", "BWP Bandwidth UL (Mhz)", "BWP Center NR-ARFCN UL", "BWP ID UL", "BWP Initial Bandwidth UL (MHz)", "BWP Initial Start Point NR-ARFCN UL", "BWP Start Point NR-ARFCN UL", "BWP Subcarrier Spacing UL",
        "BWPs Configured Count UL", "Multi_RAT_Phy_Throughput_DL_kbps", "Multi_RAT_Phy_Throughput_DL_Mbps", "Multi_RAT_Phy_Throughput_UL_kbps", "Multi RAT Connectivity Mode", "Is_Multi_RAB", "RadioAccessTechnologyState", "Handover Failure", "HandoverFailureCause", "HandoverSuccessRate", "HandoverType", "EUTRAN Inter-frequency Handover", "EUTRAN Inter-frequency Handover Failed", "EUTRAN Intra-frequency Handover", "EUTRAN Intra-frequency Handover Failed", "NR Inter-frequency Handover Attempt",
        "NR Inter-frequency Handover", "NR Inter-frequency Handover Failure", "NR Intra-frequency Handover Attempt", "NR Intra-frequency Handover"
    ]

    df_sel = df[sel_columns]

    df_sel[['Date', 'Time']] = df['Date Time'].str.split(' ', n=1, expand=True)

    df_sel.loc[:, 'CombinedColumn'] = df_sel[columns_to_combine].apply(lambda row: ' '.join(str(value) for value in row if pd.notna(value)), axis=1)
    df_sel.loc[df_sel['CombinedColumn'] == '', 'CombinedColumn'] = np.nan

    # Calculating Session ID's
    n_e_r = df_sel.index[df_sel['Ping_Address'].notna()].tolist() + [len(df_sel)]  # n_e_r = non empty rows on ping address
    labels = list(range(1, len(n_e_r) + 1)) # Numbering Session ID's
    df_sel['Session_ID'] = pd.cut(df_sel.index, bins=[0] + n_e_r, labels=labels, right=False)

    # Calculating Test ID's
    n_e_r_2 = df_sel.index[df_sel['CombinedColumn'].notna()].tolist() + [len(df_sel)]  # n_e_r_2 = non empty rows combined column
    labels_2 = list(range(1, len(n_e_r_2) + 1))  # Numbering test id's
    labels_3 = ["8.8.8.8"] + list(df_sel.loc[n_e_r_2[:-1], 'CombinedColumn'])  # First test comes from an empty row so 8.8.8.8 added
    df_sel['Test_ID'] = pd.cut(df_sel.index, bins=[0] + n_e_r_2, labels=labels_2, right=False)

    # Test Types are same as test numbering
    df_sel['Test_type'] = pd.cut(df_sel.index, bins=[0] + n_e_r_2, labels=labels_3, right=False, ordered=False)

    # Changing test type names
    my_dict = {"8.8.8.8": "PING", "URI: https://ash-speed.hetzner.com/1GB.bin": "FDTT http DL MT", "URI: http://d26kjwdsl72dzf.cloudfront.net/upload": "FDTT http UL MT", "URI: https://www.google.de/": "Google.de", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/Kepler_mobile/index.html": "Kepler", "URI: https://www.gmx.net/": "gmx.de", "URI: https://de.m.wikipedia.org/wiki/Europa": "Wikipedia", "URI: https://www.chip.de/": "Chip.de",
        "Source Uri: https://www.youtube.com/watch?v=BQwC_DJSdfE": "youtube", "URI: https://www.instagram.com/": "instagram", "URI: https://www.amazon.de/": "Amazon.de", "URI: http://212.183.159.230/5MB.zip": "FDFS http DL ST", "URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/upload/": "FDFS http UL ST", }
    df_sel['Test_type'] = df_sel['Test_type'].replace(my_dict)

    return df_sel
def calculate_general_stats(df_sel):
    from pandas import DataFrame
    Total_Stats = {}
    FDFS_DL_Attempts = df_sel['HTTP_Outcome'].count()
    Total_Stats.update({"FDFS_DL_Attempts":df_sel['HTTP_Outcome'].count()})
    FDFS_DL_Success = len(df_sel[df_sel['HTTP_Outcome'] == 'Service Status: Succeeded'])
    Total_Stats.update({"FDFS_DL_Success":FDFS_DL_Success})
    FDFS_DL_Failure = len(df_sel[df_sel['HTTP_Outcome'] == 'Service Status: Failed'])
    Total_Stats.update({"FDFS_DL_Failure":FDFS_DL_Failure})
    FDFS_DL_Success_Ratio = "{:.2%}".format(FDFS_DL_Success/FDFS_DL_Attempts)
    Total_Stats.update({"FDFS_DL_Success_Ratio":FDFS_DL_Success_Ratio})
    FDFS_UL_Attempts = df_sel['HTTP_Upload_Session_Success_Ratio'].count()
    Total_Stats.update({"FDFS_UL_Attempts":FDFS_UL_Attempts})
    FDFS_UL_Success = len(df_sel[df_sel['HTTP_Upload_Session_Success_Ratio'] == 100])
    Total_Stats.update({"FDFS_UL_Success":FDFS_UL_Success})
    FDFS_UL_Failure = len(df_sel[df_sel['HTTP_Upload_Session_Success_Ratio'] == 0])
    Total_Stats.update({"FDFS_UL_Failure":FDFS_UL_Failure})
    FDFS_UL_Success_Ratio = "{:.2%}".format(FDFS_UL_Success/FDFS_UL_Attempts)
    Total_Stats.update({"FDFS_UL_Success_Ratio":FDFS_UL_Success_Ratio})

    HTTP_Browsing_Attempts = df_sel['HTTP_Download_Session_Success_Ratio'].count()
    Total_Stats.update({'HTTP_Browsing_Attempts':HTTP_Browsing_Attempts})
    HTTP_Browsing_Access_Failure = len(df_sel[df_sel['HTTP_Download_Session_Success_Ratio'] == 0])
    Total_Stats.update({'HTTP_Browsing_Access_Failure':HTTP_Browsing_Access_Failure})
    HTTP_Browsing_Access_Success = len(df_sel[df_sel['HTTP_Download_Session_Success_Ratio'] == 100])
    Total_Stats.update({'HTTP_Browsing_Access_Success':HTTP_Browsing_Access_Success})
    HTTP_Browsing_Access_Success_Ratio = "{:.2%}".format(FDFS_DL_Success/FDFS_DL_Attempts)
    Total_Stats.update({'HTTP_Browsing_Access_Success':HTTP_Browsing_Access_Success})
    HTTP_Browsing_Min_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].min()
    Total_Stats.update({'HTTP_Browsing_Min_Transfer_Time_s':HTTP_Browsing_Min_Transfer_Time_s })
    HTTP_Browsing_10_PCTL_Transfer_Time_s = HTTP_Browsing_Min_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].quantile(0.01)
    Total_Stats.update({'HTTP_Browsing_10_PCTL_Transfer_Time_s':HTTP_Browsing_10_PCTL_Transfer_Time_s})
    HTTP_Browsing_Average_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].mean()
    Total_Stats.update({'HTTP_Browsing_Average_Transfer_Time_s':HTTP_Browsing_Average_Transfer_Time_s})
    HTTP_Browsing_Median_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].median()
    Total_Stats.update({'HTTP_Browsing_Median_Transfer_Time_s':HTTP_Browsing_Median_Transfer_Time_s})
    HTTP_Browsing_90_PCTL_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].quantile(0.9)
    Total_Stats.update({'HTTP_Browsing_90_PCTL_Transfer_Time_s':HTTP_Browsing_90_PCTL_Transfer_Time_s })
    HTTP_Browsing_Max_Transfer_Time_s = df_sel["HTTP Transfer Time (second)"].max()
    Total_Stats.update({'HTTP_Browsing_Max_Transfer_Time_s':HTTP_Browsing_Max_Transfer_Time_s})

    Video_Stream_Attempts =  df_sel['Streaming_Session_Failure_Ratio'].count()
    Video_Stream_Failures = len(df_sel[df_sel['Streaming_Session_Failure_Ratio'] == 0])
    Video_Stream_Access_Time_sec = df_sel["Streaming_Service_Access_Time_sec"].mean()
    Video_Stream_Success_Ratio_PC = len(df_sel[df_sel['Streaming_Session_Failure_Ratio'] == 100])
    Video_Stream_Reproduction_Start_Delay_sec = df_sel["Streaming_State_Request_to_Prebuffering_Delay"].mean()
    Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay = df_sel["Streaming_State_Request_to_Reproducing_Delay"].mean()
    # Streaming_Impairment_Free = df_sel["Streaming_Impairment_Free"].mean()
    Streaming_Reproduction_Cut_off_Ratio = df_sel[("Streaming_Reproduction_Cutoff_Ratio")].mean()
    Streaming_Session_Without_Interruption_Rate_PC = df_sel["Streaming_Session_Without_Interruption_Rate"].mean()
    Streaming_Aggregated_Average_Session_Resolution_p = df_sel["Streaming_Average_Session_Resolution"].mean()
    Ping_Attempt = df_sel['Ping_Count_Attempts'].count()
    Ping_Success = len(df_sel[df_sel['Ping_Count_Success'] == 10])
    Ping_Failures = len(df_sel[df_sel['Ping_Count_Failed'] == 10])
    Ping_Success_rate = "{:.2%}".format(Ping_Success/Ping_Attempt)
    Ping_RTT_ms = df_sel["Ping_Roundtrip_Time_ms"].mean()
    Ping_Trace_Loss_PC = df_sel["Ping_Packet_Loss_Rate"].mean()

    Total_Stats.update({'Video_Stream_Attempts':Video_Stream_Attempts})
    Total_Stats.update({'Video_Stream_Failures':Video_Stream_Failures})
    Total_Stats.update({'Video_Stream_Access_Time_sec':Video_Stream_Access_Time_sec})
    Total_Stats.update({'Video_Stream_Success_Ratio_PC':Video_Stream_Success_Ratio_PC})
    Total_Stats.update({'Video_Stream_Reproduction_Start_Delay_sec':Video_Stream_Reproduction_Start_Delay_sec})
    Total_Stats.update({'Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay':Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay})
    # Total_Stats.update({'Streaming_Impairment_Free_Ratio_PC':Streaming_Impairment_Free})
    Total_Stats.update({'Streaming_Reproduction_Cut-off_Ratio':Streaming_Reproduction_Cut_off_Ratio})
    Total_Stats.update({'Streaming_Session_Without_Interruption_Rate_PC':Streaming_Session_Without_Interruption_Rate_PC})
    Total_Stats.update({'Streaming_Aggregated_Average_Session_Resolution_p':Streaming_Aggregated_Average_Session_Resolution_p})
    Total_Stats.update({'Ping_Attempt':Ping_Attempt})
    Total_Stats.update({'Ping_Success':Ping_Success})
    Total_Stats.update({'Ping_Failures':Ping_Failures})
    Total_Stats.update({'Ping_Success_rate':Ping_Success_rate})

    df_total_stats = DataFrame(list(Total_Stats.items()), columns=['Statistic', 'Value'])

    return df_total_stats
    # print(df_total_stats)
def calculate_distance(row):
    from haversine import haversine, Unit
    prev_coords = (row['prev_latitude'], row['prev_longitude'])
    current_coords = (row['Latitude'], row['Longitude'])
    return haversine(prev_coords, current_coords, unit=Unit.KILOMETERS)
def add_distances(df_sel):
    df_sel['prev_latitude'] = df_sel['Latitude'].shift(1)
    df_sel['prev_longitude'] = df_sel['Longitude'].shift(1)
    df_sel['Distance'] = df_sel.apply(calculate_distance, axis=1)
    df_sel['Distance'][0] = 0
    df_sel = df_sel.drop(['prev_latitude', 'prev_longitude'], axis=1)

    # df_sel['Session_No'] = df_sel['Session_ID'].cat.codes

    return df_sel
def Calculating_pivot(df_sel):
    import pandas as pd
    import numpy as np

    df_sel['Date Time'] = pd.to_datetime(df_sel['Date Time'], format='%m/%d/%Y %H:%M:%S.%f')
    df_sel['Session_No'] = df_sel['Session_ID'].cat.codes
    # df_sel["IMSI"] = df_sel["IMSI"].fillna('').replace([np.inf, -np.inf], '').astype(str)

    agg_str = lambda x: ','.join(filter(lambda s: pd.notna(s), x))
    pcnf = lambda x: np.nanpercentile(x, 95)
    pcf = lambda x: np.nanpercentile(x, 5)
    cvah = lambda x: (x > 100000).sum()
    cvbt = lambda x: (x > 20000).sum()
    # pcnf = lambda x: np.min(x)

    agg_list = {"Latitude": ['min', 'max'], "Longitude": ['min', 'max'],
        "Date Time":['min', 'max'],
        "Date": 'min',
        "Time":'min', "Session_No": 'mean', "Test_type": agg_str, "Distance": sum, "CombinedColumn": agg_str,
        "HTTP_URL": agg_str,
        "HTTP Download Throughput (kbps)":['min','max','mean','median',pcnf,pcf,cvah,cvbt] ,
        "HTTP_Download_Service_Average_Throughput": 'mean',
        "HTTP_Download_Session_Failure_Ratio": 'mean',
        "HTTP_Download_Session_Success_Ratio": 'mean',
        "HTTP_Outcome": agg_str,
        "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A": 'mean',
        "HTTP_Download_Data_Transfer_Success_Ratio_Method_A": 'mean',
        "HTTP_Download_Data_Transfer_Time_sec_Method_A": 'mean',
        "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A": 'mean',
        "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A": 'mean',
        "HTTP_Download_IP_Service_Setup_Time_sec_Method_A": 'mean',
        "HTTP_Download_Mean_Data_Rate_kbps_Method_A": 'mean',
        "HTTP_Download_Transfer_Start_Delay_Method_A": 'mean',
        "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B": 'mean',
        "HTTP_Download_Data_Transfer_Success_Ratio_Method_B": 'mean',
        "HTTP_Download_Data_Transfer_Time_sec_Method_B": 'mean',
        "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B": 'mean',
        "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B": 'mean',
        "HTTP_Download_IP_Service_Setup_Time_sec_Method_B": 'mean',
        "HTTP_Download_Mean_Data_Rate_kbps_Method_B": 'mean',
        "HTTP_Upload_Average_Throughput": 'mean', "HTTP_Upload_Service_Average_Throughput": 'mean',
        "HTTP_Upload_Service_Transfer_Time": 'mean', "HTTP_Upload_Session_Failure_Ratio": 'mean',
        "HTTP_Upload_Session_Success_Ratio": 'mean', "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A": 'mean', "HTTP_Upload_Data_Transfer_Time_sec_Method_A": 'mean', "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A": 'mean', "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A": 'mean',
        "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A": 'mean', "HTTP_Upload_Mean_Data_Rate_kbps_Method_A": 'mean',
        "Streaming_URL": agg_str, "Streaming_Outcome_Type": agg_str, "Aborted_by_User": 'mean',
        "Streaming_Average_Session_Resolution": 'mean',
        "Streaming_Average_Throughput": 'mean',
        "Streaming_Completion_Rate": 'mean',
        "Streaming_Duration": 'mean',
        "Streaming_HD_Resolution": agg_str, "Streaming_HD_Resolution_Ratio": 'mean', "Streaming_Impairment_Free": agg_str,
        "Streaming_Impairment_Free_Video_Session_Ratio": 'mean',
        "Streaming_Maximum_Duration_Of_Video_Session_Interruptions": 'mean',
        "Streaming_Number_Of_Video_Session_Interruptions": 'mean', "Streaming_Player_Size_kB": 'mean',
        "Streaming_Reproduction_Cutoff_Ratio": 'mean',
        "Streaming_Reproduction_Start_Failure_Ratio": 'mean',
        "Streaming_Reproduction_Start_Delay_sec":'mean',
        "Streaming_State_Request_to_Prebuffering_Delay":'mean',
        "Streaming_Service_Access_Time_ms": 'mean',
        "Streaming_Service_Access_Time_sec": 'mean',
        "Streaming_Session_Failure_Ratio": 'mean',
        "Streaming_Session_Qualified": agg_str,
        "Streaming_Session_Qualified_Ratio": 'mean',
        "Streaming_Session_Video_Interruption_Duration": 'mean',
        "Streaming_Session_Without_Interruption_Rate": 'mean',
        "Streaming_Setup_Success_Rate": 'mean',
        "Streaming_State_Prebuffering_to_Reproducing_Delay": 'mean',
        "Streaming_State_Request_to_Reproducing_Delay": 'mean',
        "Streaming_Success_Rate": 'mean',
        "Streaming_Throughput_Filtered": 'mean',
        "Streaming_Total_Duration_Of_Video_Session_Interruptions": 'mean',
        "Streaming_Video_Buffer_Size_kB": 'mean',
        "Streaming_Video_Play_Start_Failure_Ratio": 'mean',
        "Streaming_Video_Play_Start_Time_sec": 'mean',
        "Streaming_Video_Session_Cutoff_Ratio": 'mean',
        "Streaming_Video_Session_Failure_Ratio": 'mean',
        "Streaming_Video_Session_Success_Ratio": 'mean',
        "Streaming_Video_Session_Time_sec": 'mean',
        "Streaming_Video_Size_kB": 'mean',
        #"SessionTime": 'mean',
        #"SessionTime_Upload": 'mean',
        "ThroughputCountOver1MBit": 'mean', "ThroughputCountOver3MBit": 'mean', "ThroughputPercentageOver3MBit": 'mean',
        "TCP_Handshake_Time_sec": 'mean', "Ping_Address": agg_str,
        "Ping_Count_Attempts": 'mean',
        "Ping_Count_Failed": 'mean',
        "Ping_Count_Success": 'mean',
        "Ping_Delay_ms_Avg": 'mean', "Ping_Delay_ms_Max": 'mean', "Ping_Delay_ms_Min": 'mean',
        "Ping_Packet_Loss_Rate": 'mean', "Ping_Packet_Success_Rate": 'mean',
        "Ping_Roundtrip_Time_ms": 'mean',
        "Ping_Roundtrip_Time_sec": 'mean',
        "Ping_Size": 'mean',
        #"Data_Radio_Bearer": agg_str,
        "Fixed_Duration": 'mean', "IP_Interruption_Time_ms": 'mean', "Is_Multi_RAB": agg_str, "LTE_Serving_Cell_Count_Average": 'mean', "Service_Bearer": agg_str, "DNS_Client": agg_str, "DNS_Domain_Name": agg_str, "DNS_First_In_Session": agg_str, "DNS_Host_Name_Resolution_Failure_Ratio": 'mean', "DNS_Host_Name_Resolution_Time_sec": 'mean',
        "DNS_Host_Name_Total_Resolution_Time_sec": 'mean', "DNS_Resolved_Address": agg_str,
        "DNS_Server_Address": agg_str,
        # "PCell_RAT":agg_str,
        # "Campaign" : agg_str,
        # "Operator" : agg_str,
        "RadioAccessTechnologyState": 'mean',
        "IMEI":agg_str,
        "IMSI":'mean'
    }

    pivot_result = pd.pivot_table(df_sel, index=["Test_ID"], values=list(agg_list.keys()), aggfunc=agg_list)

    pivot_result.columns = [' '.join(map(str, tpl)) for tpl in pivot_result.columns]
    pivot_result.columns = [s.replace(" <lambda>", '') for s in pivot_result.columns]
    pivot_result.columns = [s.replace(" mean", '') for s in pivot_result.columns]
    df_pivot = pivot_result.reset_index()

    # print(df_pivot)

    def clear_duplicate_values_in_cell(cell_value):
        values = cell_value.split(',')
        unique_values = list(set(values))
        return ','.join(unique_values)
    def clear_duplicate_values_in_column(dataframe, column_name):
        df_copy = dataframe.copy()
        df_copy[column_name] = df_copy[column_name].apply(clear_duplicate_values_in_cell)
        return df_copy

    df_result = clear_duplicate_values_in_column(df_pivot, "DNS_Client")
    df_result = clear_duplicate_values_in_column(df_result, "DNS_Domain_Name")
    df_result = clear_duplicate_values_in_column(df_result, "DNS_Resolved_Address")
    df_result = clear_duplicate_values_in_column(df_result, "DNS_Server_Address")
    df_result = clear_duplicate_values_in_column(df_result, "CombinedColumn")
    df_result = clear_duplicate_values_in_column(df_result, "Test_type")
    df_result = clear_duplicate_values_in_column(df_result, "IMEI")
    df_result = clear_duplicate_values_in_column(df_result, "Is_Multi_RAB")

    # print(df_result["IMSI"].head(30))

    df_result = df_result.rename(columns={'CombinedColumn': 'Grouping_All_URL'})
    df_result["Duration"] = df_result["Date Time max"] - df_result["Date Time min"]
    # print(df_result["Duration"].head(10))
    df_result['Duration_Seconds'] = df_result['Duration'].apply(lambda x: pd.Timedelta(x).total_seconds())

    return df_result
def merge_type_table(df_result, test_code,operator):
    import pandas as pd
    Grouping_All_Url = [
"8.8.8.8",
"Source Uri: https://www.youtube.com/watch?v=BQwC_DJSdfE",
"URI: http://212.183.159.230/5MB.zip",
"URI: http://d26kjwdsl72dzf.cloudfront.net/upload",
"URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/Kepler_mobile/index.html",
"URI: http://press21deq5.s3.dualstack.eu-central-1.amazonaws.com/upload/",
"URI: https://ash-speed.hetzner.com/1GB.bin",
"URI: https://de.m.wikipedia.org/wiki/Europa",
"URI: https://www.amazon.de/",
"URI: https://www.chip.de/",
"URI: https://www.gmx.net/",
"URI: https://www.google.de/",
"URI: https://www.instagram.com/"]
    Direction_Test = [
"PING",
"DL",
"DL",
"UL",
"DL",
"UL",
"DL",
"DL",
"DL",
"DL",
"DL",
"DL",
"DL"]
    Test_Name = [
"PING",
"Streaming YT",
"FDFS HTTP DL ST",
"FDTT HTTP UL MT",
"Kepler",
"FDFS HTTP UL ST",
"FDTT HTTP DL MT",
"Wikipedia",
"Amazon.de",
"Chip.de",
"Gmx.net",
"Google.de",
"Instagram.com"]
    Type_of_Test = [
"PING",
"Streaming YT",
"FDFS DL",
"FDTT UL ",
"HTTP Browsing",
"FDFS UL",
"FDTT DL",
"HTTP Browsing",
"HTTP Browsing",
"HTTP Browsing",
"HTTP Browsing",
"HTTP Browsing",
"HTTP Browsing"]
    Thread_info = [
"PING",
"Live 4K",
"ST",
"MT",
"MT",
"ST",
"MT",
"MT",
"MT",
"MT",
"MT",
"MT",
"MT"
]

    data = {
    'Grouping_All_URL': Grouping_All_Url ,
    'Direction_Test': Direction_Test,
    'Test_Name':Test_Name,
    'Type_of_Test':Type_of_Test,
    'Thread_info':Thread_info }
    df_gr = pd.DataFrame(data)


    merged_df = pd.merge(df_result, df_gr, on='Grouping_All_URL', how='left')

    columns_to_move = ['Grouping_All_URL','Direction_Test', 'Test_Name','Type_of_Test','Thread_info']
    new_positions = [3,4,5,6,7]  # 0-indexed positions

    # Move each column to its new position
    for column_name, new_position in zip(columns_to_move, new_positions):
        column_to_move = merged_df.pop(column_name)
        merged_df.insert(new_position, column_name, column_to_move)

    # merged_df["Year"] = merged_df["Date min"].str[:-4]
    # merged_df["Month"] = merged_df["Date min"].str[5:7]
    # merged_df["Day"] = merged_df["Date min"].str[8:10]
    # merged_df["Hour"] = merged_df["Time min"].str[:2]

    merged_df[['Day','Month','Year']] = merged_df['Date min'].str.split('/', n=3, expand=True)
    merged_df["Hour"] = merged_df["Time min"].str[:2]

    merged_df["Sequenz_ID_per_File"] = merged_df["Test_ID"]
    # merged_df["Test_ID"] =  merged_df['Year'] + merged_df['Month'] +  merged_df['Day'] + merged_df["Test_ID"].astype(str)

    merged_df["Test_ID"] =  merged_df['Year'] + test_code + merged_df['Month'].str.zfill(2) +merged_df['Day'].str.zfill(2) + merged_df["Test_ID"].astype(str).str.zfill(4)

    merged_df["Operator"] = operator
    merged_df["Campaign"] = " "
    Col_ordered = [
    "Campaign",
    "Operator",
    "Type_of_Test",
    "Technology",
    "Test_Name",
    "Direction_Test",
    "Thread_info",
    "Grouping_All_URL",
    "Test_ID",
    "Session_No",
    "Sequenz_ID_per_File",
    "Duration_Seconds",
    "Day",
    "Month",
    "Year",
    "Hour",
    "Distance sum",
    "TCP_Handshake_Time_sec",
    "Is_Multi_RAB",
    "IP_Interruption_Time_ms",
    "ThroughputCountOver1MBit",
    "ThroughputCountOver3MBit",
    "ThroughputPercentageOver3MBit",
    "RadioAccessTechnologyState",
    "Service_Bearer",
    "LTE_Serving_Cell_Count_Average",
    "DNS_Client",
    "DNS_Domain_Name",
    "DNS_First_In_Session",
    "DNS_Host_Name_Resolution_Failure_Ratio",
    "DNS_Host_Name_Resolution_Time_sec",
    "DNS_Host_Name_Total_Resolution_Time_sec",
    "DNS_Resolved_Address",
    "DNS_Server_Address",
    "HTTP Download Throughput (kbps) <lambda_0>",
    "HTTP Download Throughput (kbps) <lambda_1>",
    "HTTP Download Throughput (kbps) <lambda_2>",
    "HTTP Download Throughput (kbps) <lambda_3>",
    "HTTP Download Throughput (kbps) max",
    "HTTP Download Throughput (kbps)",
    "HTTP Download Throughput (kbps) median",
    "HTTP Download Throughput (kbps) min",
    "HTTP_Download_Data_Transfer_Failure_Ratio_Method_A",
    "HTTP_Download_Data_Transfer_Failure_Ratio_Method_B",
    "HTTP_Download_Data_Transfer_Success_Ratio_Method_A",
    "HTTP_Download_Data_Transfer_Success_Ratio_Method_B",
    "HTTP_Download_Data_Transfer_Time_sec_Method_A",
    "HTTP_Download_Data_Transfer_Time_sec_Method_B",
    "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_A",
    "HTTP_Download_IP_Service_Access_Failure_Ratio_Method_B",
    "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_A",
    "HTTP_Download_IP_Service_Setup_Success_Ratio_Method_B",
    "HTTP_Download_IP_Service_Setup_Time_sec_Method_A",
    "HTTP_Download_IP_Service_Setup_Time_sec_Method_B",
    "HTTP_Download_Mean_Data_Rate_kbps_Method_A",
    "HTTP_Download_Mean_Data_Rate_kbps_Method_B",
    "HTTP_Download_Service_Average_Throughput",
    "HTTP_Download_Session_Failure_Ratio",
    "HTTP_Download_Session_Success_Ratio",
    "HTTP_Download_Transfer_Start_Delay_Method_A",
    "HTTP_Outcome",
    "HTTP_URL",
    "HTTP_Upload_Average_Throughput",
    "HTTP_Upload_Data_Transfer_Failure_Ratio_Method_A",
    "HTTP_Upload_Data_Transfer_Success_Ratio_Method_A",
    "HTTP_Upload_Data_Transfer_Time_sec_Method_A",
    "HTTP_Upload_IP_Service_Access_Failure_Ratio_Method_A",
    "HTTP_Upload_IP_Service_Setup_Success_Ratio_Method_A",
    "HTTP_Upload_IP_Service_Setup_Time_sec_Method_A",
    "HTTP_Upload_Mean_Data_Rate_kbps_Method_A",
    "HTTP_Upload_Service_Average_Throughput",
    "HTTP_Upload_Service_Transfer_Time",
    "HTTP_Upload_Session_Failure_Ratio",
    "HTTP_Upload_Session_Success_Ratio",
    "Ping_Address",
    "Ping_Count_Attempts",
    "Ping_Count_Failed",
    "Ping_Count_Success",
    "Ping_Delay_ms_Avg",
    "Ping_Delay_ms_Max",
    "Ping_Delay_ms_Min",
    "Ping_Packet_Loss_Rate",
    "Ping_Packet_Success_Rate",
    "Ping_Roundtrip_Time_ms",
    "Ping_Roundtrip_Time_sec",
    "Ping_Size",
    "Streaming_Average_Session_Resolution",
    "Streaming_Average_Throughput",
    "Streaming_Completion_Rate",
    "Streaming_Duration",
    "Streaming_HD_Resolution",
    "Streaming_HD_Resolution_Ratio",
    "Streaming_Impairment_Free",
    "Streaming_Impairment_Free_Video_Session_Ratio",
    "Streaming_Maximum_Duration_Of_Video_Session_Interruptions",
    "Streaming_Number_Of_Video_Session_Interruptions",
    "Streaming_Outcome_Type",
    "Streaming_Reproduction_Cutoff_Ratio",
    "Streaming_Reproduction_Start_Delay_sec",
    "Streaming_Reproduction_Start_Failure_Ratio",
    "Streaming_Service_Access_Time_ms",
    "Streaming_Service_Access_Time_sec",
    "Streaming_Session_Failure_Ratio",
    "Streaming_Session_Qualified",
    "Streaming_Session_Without_Interruption_Rate",
    "Streaming_Setup_Success_Rate",
    "Streaming_State_Prebuffering_to_Reproducing_Delay",
    "Streaming_State_Request_to_Prebuffering_Delay",
    "Streaming_State_Request_to_Reproducing_Delay",
    "Streaming_Success_Rate",
    "Streaming_Throughput_Filtered",
    "Streaming_Total_Duration_Of_Video_Session_Interruptions",
    "Streaming_URL",
    "Streaming_Video_Play_Start_Failure_Ratio",
    "Streaming_Video_Play_Start_Time_sec",
    "Streaming_Video_Session_Cutoff_Ratio",
    "Streaming_Video_Session_Failure_Ratio",
    "Streaming_Video_Session_Success_Ratio",
    "Streaming_Video_Session_Time_sec",
    "IMEI",
    "IMSI",
    ]

    if 'IMSI' not in merged_df.columns:
        merged_df['IMSI'] = ''
    merged_df["Technology"] = merged_df["RadioAccessTechnologyState"]
    mapping = {2: 'GSM/WCDMA', 3: 'LTE', 7: 'EN-DC'}
    merged_df["Technology"] =  merged_df["Technology"].map(mapping).fillna('LTE/EN-DC')

    df_ordered = merged_df[Col_ordered]

    return df_ordered
def data_preprocessing_voice(df):
    import pandas as pd
    import numpy as np

    columns_to_combine = ["HTTP_URL", "Streaming_URL", "Ping_Address"]

    sel_columns = [
    "Time",
    "Date",
    "Latitude",
    "Longitude",
    "LogfileName",
    "EventName",
    "DeviceDescription",
    "IMEI",
    "IMSI",
    "MCC",
    "MNC",
    "SIM Operator",
    "Technology_Detail",
    "Dialed Number",
    "A_Number",
    "B_Number",
    "Call_Status",
    "Call_Type",
    "Call_Number",
    "MO_Call_Domain",
    "MO_Call_Setup_Time_ms",
    "MO_Call_User_Setup_Time_ms",
    "MOCallDropRate",
    "MOCallSetupFailureRate",
    "Call_Duration",
    "Call_Setup_Duration",
    "Call_Setup_Time_sec",
    "Voice_Call_Setup_Time_sec",
    "Voice_Call_Success_Rate",
    "Voice_Call_Setup_Success_Rate",
    "Voice_Call_Setup_Failure_Rate",
    "Voice_Call_Completion_Rate",
    "Voice_Dropped_Call_Rate",
    "Voice_Speech_Quality_On_Call_Basis",
    "Call Drop Rate",
    "Call Dropped Count",
    "Call Dropped VoLTE Count",
    "Call Dropped VoNR Count",
    "Call Block Rate",
    "Call Blocked Count",
    "Call Blocked VoLTE Count",
    "Call Blocked VoNR Count",
    "Call Success Count",
    "Call Success Rate",
    "Call Success VoLTE Count",
    "Call Success VoNR Count",
    "Call Initiation",
    "Call Attempt",
    "Call Attempt Retry",
    "Call Setup",
    "Call Established",
    "Call End",
    "Call_End_Initiator",
    "Data Radio Bearer",
    "Multi RAT Connectivity Mode",
    "RadioAccessTechnologyState",
    "Serving Cell DL EARFCN",
    "Serving Cell eMTC Cell Indication",
    "Serving Cell Frequency Band",
    "Serving Cell Identity",
    "Serving Cell RSRP (dBm)",
    "Serving Cell RSRQ (dB)",
    "Serving Cell RS SINR (dB)",
    "Serving Cell Serving Cell EN-DC Support",
    "MR-DC Cell 1 Band",
    "MR-DC Cell 1 Channel",
    "MR-DC Cell 1 PCI",
    "MR-DC Cell 1 RSRP (dBm)",
    "MR-DC Cell 1 RSRQ (dB)",
    "MR-DC Cell 1 SINR (dB)",
    "CallDroppedRate",
    "DRBDroppedRate",
    "SipByeReasonPhrase",
    "SpeechPathDelayRTT",
    "SpeechPathDelayOneWay",
    "RTPInterruptionTimeAudio",
    "AQMScoreDL",
    "AQMAlgorithmDL",
    "AQM_RAT",
    "AQM Algorithm DL",
    "AQM Audio Channel Type DL",
    "AQM Score",
    "AQM Score DL",
    "MOS Calculator Version Downlink",
    "Speech Codec",
    "AMR_Codec_DL",
    "SQ_MOS",
    "MultiRABServiceType1",
    "MultiRABServiceType2",
    "AQMDownlinkCallQuality",
    "AverageAQMDownlinkCallQuality",
    "POLQA SWB Score DL",
    "Speech Codec",
    "RTP Jitter Audio RFC3550 (ms)",
    "RTP Lost Packets Rate Audio (%)",
    "RTP Relative Packet Delay Audio (ms)",
    "Application Layer kBytes Received (Cumulative)",
    "Application Layer kBytes Received (Current)",
    "HTTP_Download_Average_Throughput",
    "HTTP_Download_Service_Average_Throughput",
    "HTTP_Download_Session_Success_Ratio",
    "HTTP_Download_Session_Failure_Ratio",
    "HTTP_Outcome",
    "HTTP_URL",
    "HTTP_Download_Bearer",
    "HTTP Transfer Time (second)",
    ]

    df_sel = df[sel_columns]
    df_sel.rename(columns={'Date': 'Date Time'}, inplace=True)

    df_sel[['Date', 'Time']] = df_sel['Date Time'].str.split(' ', n=1, expand=True)

    df_sel.loc[df_sel['EventName'] == '', 'EventName'] = np.nan

    n_e_r_2 = df_sel.index[df_sel['EventName']=="MO Session Attempt"].tolist() + [len(df_sel)]  # n_e_r_2 = non empty rows combined column
    labels_2 = list(range(0, len(n_e_r_2)))  # Numbering test id's

    df_sel['Test_ID'] = pd.cut(df_sel.index, bins=[1-0] + n_e_r_2, labels=labels_2, right=False)

    # print(df_sel)
    return df_sel
def Calculating_pivot_voice(df_sel,test_code):
    import pandas as pd
    import numpy as np

    df_sel['Date Time'] = pd.to_datetime(df_sel['Date Time'], format='%m/%d/%Y %H:%M:%S.%f')
    # df_sel['Session_No'] = df_sel['Session_ID'].cat.codes

    agg_str = lambda x: ','.join(filter(lambda s: pd.notna(s), x))
    pcnf = lambda x: np.nanpercentile(x, 95)
    pcf = lambda x: np.nanpercentile(x, 5)
    cvah = lambda x: (x > 100000).sum()
    cvbt = lambda x: (x > 20000).sum()


    df_sel["Serving Cell Identity"] = df_sel["Serving Cell Identity"].fillna('').replace([np.inf, -np.inf], '').astype(str)
    df_sel["Serving Cell DL EARFCN"] = df_sel["Serving Cell DL EARFCN"].fillna('').replace([np.inf, -np.inf], '').astype(str)

    agg_list ={
    "Date":['min', 'max'],
    "Time":min,
    "Date Time":['min', 'max'],
    "Latitude":['min', 'max'],
    "Longitude":['min', 'max'],
    "IMEI":agg_str,
    "IMSI":'mean',
    "MCC":'mean',
    "MNC":'mean',
    "SIM Operator":agg_str,
    "Technology_Detail":agg_str,
    "MO_Call_Setup_Time_ms":'mean',
    "MO_Call_User_Setup_Time_ms":'mean',
    "MOCallDropRate":'mean',
    "MOCallSetupFailureRate":'mean',
    "Call_Duration":'mean',
    "Call_Setup_Duration":'mean',
    "Call_Setup_Time_sec":'mean',
    "Voice_Call_Setup_Time_sec":'mean',
    "Voice_Call_Success_Rate":'mean',
    "Voice_Call_Setup_Success_Rate":'mean',
    "Voice_Call_Setup_Failure_Rate":'mean',
    "Voice_Call_Completion_Rate":'mean',
    "Voice_Dropped_Call_Rate":'mean',
    "Voice_Speech_Quality_On_Call_Basis":'mean',
    "Call Drop Rate":'mean',
    "Call Dropped Count":'mean',
    "Call Dropped VoLTE Count":'mean',
    "Call Dropped VoNR Count":'mean',
    "Call Block Rate":'mean',
    "Call Blocked Count":'mean',
    "Call Blocked VoLTE Count":'mean',
    "Call Blocked VoNR Count":'mean',
    "Call Success Count":'mean',
    "Call Success Rate":'mean',
    "Call Success VoLTE Count":'mean',
    "Call Success VoNR Count":'mean',
    "Data Radio Bearer":agg_str,
    "Multi RAT Connectivity Mode":agg_str,
    "RadioAccessTechnologyState":'mean',
    "Serving Cell DL EARFCN":agg_str,
    "Serving Cell eMTC Cell Indication":'mean',
    "Serving Cell Frequency Band": agg_str,
    "Serving Cell Identity":agg_str,
    "Serving Cell RSRP (dBm)":'mean',
    "Serving Cell RSRQ (dB)":'mean',
    "Serving Cell RS SINR (dB)":'mean',
    "Serving Cell Serving Cell EN-DC Support":'mean',
    "CallDroppedRate":'mean',
    "DRBDroppedRate":'mean',
    "SipByeReasonPhrase": agg_str,
    "SpeechPathDelayRTT":'mean',
    "SpeechPathDelayOneWay":'mean',
    "RTPInterruptionTimeAudio":'mean',
    "AQMScoreDL":'mean',
    "AQMAlgorithmDL":agg_str,
    "AQM_RAT":agg_str,
    "AQM Algorithm DL":agg_str,
    "AQM Audio Channel Type DL":'mean',
    "AQM Score":'mean',
    "AQM Score DL":'mean',
    "MOS Calculator Version Downlink":agg_str,
    # "Speech Codec":agg_str,
    "AMR_Codec_DL":agg_str,
    "SQ_MOS":'mean',
    "RTP Jitter Audio RFC3550 (ms)":'mean',
    "RTP Lost Packets Rate Audio (%)":'mean',
    "RTP Relative Packet Delay Audio (ms)":'mean',
    "Application Layer kBytes Received (Cumulative)":'mean',
    "Application Layer kBytes Received (Current)":'mean',
    "HTTP_Download_Average_Throughput":'mean',
    "HTTP_Download_Service_Average_Throughput":'mean',
    "HTTP_Download_Session_Success_Ratio":'mean',
    "HTTP_Download_Session_Failure_Ratio":'mean',
    "HTTP_Outcome":agg_str,
    "HTTP_URL":agg_str,
    "HTTP_Download_Bearer":agg_str,
    "HTTP Transfer Time (second)":'mean',
    "Distance": sum,
    "POLQA SWB Score DL":'mean',
    "Call_Status" : agg_str,
    }

    pivot_result = pd.pivot_table(df_sel, index=["Test_ID"], values=list(agg_list.keys()), aggfunc=agg_list)

    pivot_result.columns = [' '.join(map(str, tpl)) for tpl in pivot_result.columns]
    pivot_result.columns = [s.replace(" <lambda>", '') for s in pivot_result.columns]
    pivot_result.columns = [s.replace(" mean", '') for s in pivot_result.columns]
    df_pivot = pivot_result.reset_index()

    def clear_duplicate_values_in_cell(cell_value):
        values = cell_value.split(',')
        unique_values = list(set(values))
        return ','.join(unique_values)
    def clear_duplicate_values_in_column(dataframe, column_name):
        df_copy = dataframe.copy()
        df_copy[column_name] = df_copy[column_name].apply(clear_duplicate_values_in_cell)
        return df_copy

    agg_clear_list = [
    "Call_Status",
    "IMEI",
    "SIM Operator",
    "Technology_Detail",
    "Data Radio Bearer",
    "Multi RAT Connectivity Mode",
    "Serving Cell Frequency Band",
    "Serving Cell Identity",
    "SipByeReasonPhrase",
    "AQMAlgorithmDL",
    "AQM_RAT",
    "AQM Algorithm DL",
    "MOS Calculator Version Downlink",
    "AMR_Codec_DL",
    "HTTP_Outcome",
    "HTTP_URL",
    "HTTP_Download_Bearer",
    "Serving Cell DL EARFCN"
    ]

    for i in agg_clear_list:
        df_pivot = clear_duplicate_values_in_column(df_pivot,i)

    # df_pivot = clear_duplicate_values_in_column(df_pivot, "DNS_Resolved_Address")
    # df_result = clear_duplicate_values_in_column(df_result, "DNS_Server_Address")
    # df_result = clear_duplicate_values_in_column(df_result, "CombinedColumn")
    # df_result = clear_duplicate_values_in_column(df_result, "Test_type")

    # df_result = df_result.rename(columns={'CombinedColumn': 'Grouping_All_URL'})
    df_pivot["Duration"] = df_pivot["Date Time max"] - df_pivot["Date Time min"]
    # print(df_result["Duration"].head(10))
    df_pivot['Duration_Seconds'] = df_pivot['Duration'].apply(lambda x: pd.Timedelta(x).total_seconds())

    df_pivot[['Month','Day','Year']] = df_pivot['Date min'].str.split('/', n=3, expand=True)
    df_pivot["Hour"] = df_pivot["Time min"].str[:2]

    df_pivot = df_pivot[df_pivot['Test_ID'] != 0]
    df_pivot["Sequenz_ID_per_File"] = df_pivot["Test_ID"]
    df_pivot["Test_ID"] =  df_pivot['Year'] +test_code +df_pivot['Month'].str.zfill(2) +df_pivot['Day'].str.zfill(2) + df_pivot["Test_ID"].astype(str).str.zfill(4)

    col_arrange = ["Test_ID",
    "Sequenz_ID_per_File",
    "Day",
    "Month",
    "Year",
    "Hour",
    "Duration_Seconds",
    "IMEI",
    "IMSI",
    "MCC",
    "MNC",
    "Multi RAT Connectivity Mode",
    "RadioAccessTechnologyState",
    "SIM Operator",
    "Distance sum",
    "SipByeReasonPhrase",
    "Technology_Detail",
    "MOCallDropRate",
    "MOCallSetupFailureRate",
    "MOS Calculator Version Downlink",
    "MO_Call_Setup_Time_ms",
    "MO_Call_User_Setup_Time_ms",
    "Call Block Rate",
    "Call Blocked Count",
    "Call Blocked VoLTE Count",
    "Call Blocked VoNR Count",
    "Call Drop Rate",
    "Call Dropped Count",
    "Call Dropped VoLTE Count",
    "Call Dropped VoNR Count",
    "Call Success Count",
    "Call Success Rate",
    "Call Success VoLTE Count",
    "Call Success VoNR Count",
    "CallDroppedRate",
    "Call_Duration",
    "Call_Setup_Duration",
    "Call_Setup_Time_sec",
    "Call_Status",
    "Voice_Call_Completion_Rate",
    "Voice_Call_Setup_Failure_Rate",
    "Voice_Call_Setup_Success_Rate",
    "Voice_Call_Setup_Time_sec",
    "Voice_Call_Success_Rate",
    "Voice_Dropped_Call_Rate",
    "Voice_Speech_Quality_On_Call_Basis",
    "SQ_MOS",
    "POLQA SWB Score DL",
    "AQM Algorithm DL",
    "AQM Audio Channel Type DL",
    "AQM Score",
    "AQM Score DL",
    "AQMAlgorithmDL",
    "AQMScoreDL",
    "AQM_RAT",
    "RTP Jitter Audio RFC3550 (ms)",
    "RTP Lost Packets Rate Audio (%)",
    "RTP Relative Packet Delay Audio (ms)",
    "RTPInterruptionTimeAudio",
    "HTTP Transfer Time (second)",
    "HTTP_Download_Average_Throughput",
    "HTTP_Download_Bearer",
    "HTTP_Download_Service_Average_Throughput",
    "HTTP_Download_Session_Failure_Ratio",
    "HTTP_Download_Session_Success_Ratio",
    "HTTP_Outcome",
    "HTTP_URL",
    "DRBDroppedRate",
    "Data Radio Bearer",
    "Application Layer kBytes Received (Cumulative)",
    "Application Layer kBytes Received (Current)",
    "AMR_Codec_DL",
    "Serving Cell DL EARFCN",
    "Serving Cell Frequency Band",
    "Serving Cell Identity",
    "Serving Cell RS SINR (dB)",
    "Serving Cell RSRP (dBm)",
    "Serving Cell RSRQ (dB)",
    "Serving Cell Serving Cell EN-DC Support",
    "Serving Cell eMTC Cell Indication"]



    df_pivot['Serving Cell Identity'] = "'" + df_pivot['Serving Cell Identity'].str.lstrip(',').str.split(',').apply(lambda x: ','.join([str(int(float(x[0])))] + [str(int(float(value))) if '.' in value else str(int(value)) for value in x[1:]]))+ "'"
    df_pivot["Serving Cell DL EARFCN"] = "'" + df_pivot['Serving Cell DL EARFCN'].str.lstrip(',').str.split(',').apply(lambda x: ','.join([str(int(float(x[0])))] + [str(int(float(value))) if '.' in value else str(int(value)) for value in x[1:]]))+ "'"

    df_ordered = df_pivot[col_arrange]

    print( df_ordered["Serving Cell DL EARFCN"])

    return df_ordered
def data_preprocessing_voice_MT(df):
    import pandas as pd
    import numpy as np


    sel_columns = [
    "Time",
    "Date",
    "Latitude",
    "Longitude",
    "LogfileName",
    "EventName",
    "DeviceDescription",
    "IMEI",
    "IMSI",
    "MCC",
    "MNC",
    "SIM Operator",
    "Technology_Detail",
    "Dialed Number",
    "A_Number",
    "B_Number",
    "Call_Status",
    "Call_Type",
    "Call_Number",
    "MT_Call_Domain",
    "MT_Call_Setup_Time_ms",
    "Call_Duration",
    "Call_Setup_Duration",
    "Call_Setup_Time_sec",
    "Voice_Call_Setup_Time_sec",
    "Voice_Call_Success_Rate",
    "Voice_Call_Setup_Success_Rate",
    "Voice_Call_Setup_Failure_Rate",
    "Voice_Call_Completion_Rate",
    "Voice_Dropped_Call_Rate",
    "Voice_Speech_Quality_On_Call_Basis",
    "Call Drop Rate",
    "Call Dropped Count",
    "Call Dropped VoLTE Count",
    "Call Dropped VoNR Count",
    "Call Block Rate",
    "Call Blocked Count",
    "Call Blocked VoLTE Count",
    "Call Blocked VoNR Count",
    "Call Success Count",
    "Call Success Rate",
    "Call Success VoLTE Count",
    "Call Success VoNR Count",
    "Call Initiation",
    "Call Attempt",
    "Call Attempt Retry",
    "Call Setup",
    "Call Established",
    "Call End",
    "Call_End_Initiator",
    "Data Radio Bearer",
    "Multi RAT Connectivity Mode",
    "RadioAccessTechnologyState",
    "Serving Cell DL EARFCN",
    "Serving Cell eMTC Cell Indication",
    "Serving Cell Frequency Band",
    "Serving Cell Identity",
    "Serving Cell RSRP (dBm)",
    "Serving Cell RSRQ (dB)",
    "Serving Cell RS SINR (dB)",
    "Serving Cell Serving Cell EN-DC Support",
    "MR-DC Cell 1 Band",
    "MR-DC Cell 1 Channel",
    "MR-DC Cell 1 PCI",
    "MR-DC Cell 1 RSRP (dBm)",
    "MR-DC Cell 1 RSRQ (dB)",
    "MR-DC Cell 1 SINR (dB)",
    "CallDroppedRate",
    "DRBDroppedRate",
    "SipByeReasonPhrase",
    "SpeechPathDelayRTT",
    "SpeechPathDelayOneWay",
    "RTPInterruptionTimeAudio",
    "AQMScoreDL",
    "AQMAlgorithmDL",
    "AQM_RAT",
    "AQM Algorithm DL",
    "AQM Audio Channel Type DL",
    "AQM Score",
    "AQM Score DL",
    "MOS Calculator Version Downlink",
    "Speech Codec",
    "AMR_Codec_DL",
    "SQ_MOS",
    "MultiRABServiceType1",
    "MultiRABServiceType2",
    "AQMDownlinkCallQuality",
    "AverageAQMDownlinkCallQuality",
    "POLQA SWB Score DL",
    "Speech Codec",
    "RTP Jitter Audio RFC3550 (ms)",
    "RTP Lost Packets Rate Audio (%)",
    "RTP Relative Packet Delay Audio (ms)",
    "Application Layer kBytes Received (Cumulative)",
    "Application Layer kBytes Received (Current)",
    "HTTP_Download_Average_Throughput",
    "HTTP_Download_Service_Average_Throughput",
    "HTTP_Download_Session_Success_Ratio",
    "HTTP_Download_Session_Failure_Ratio",
    "HTTP_Outcome",
    "HTTP_URL",
    "HTTP_Download_Bearer",
    "HTTP Transfer Time (second)"
    ]

    df_sel = df[sel_columns]
    df_sel.rename(columns={'Date': 'Date Time'}, inplace=True)

    df_sel[['Date', 'Time']] = df_sel['Date Time'].str.split(' ', n=1, expand=True)

    # df_sel.loc[:, 'CombinedColumn'] = df_sel[columns_to_combine].apply(lambda row: ' '.join(str(value) for value in row if pd.notna(value)), axis=1)
    df_sel.loc[df_sel['EventName'] == '', 'EventName'] = np.nan

    # Calculating Test ID's
    n_e_r_2 = df_sel.index[df_sel['EventName']=="MT Session Attempt"].tolist() + [len(df_sel)]  # n_e_r_2 = non empty rows combined column
    labels_2 = list(range(0, len(n_e_r_2) ))  # Numbering test id's
    df_sel['Test_ID'] = pd.cut(df_sel.index, bins=[1] + n_e_r_2, labels=labels_2, right=False)
    # df_sel = df_sel[df_sel['Test_ID'] != '0']

    return df_sel
def Calculating_pivot_voice_MT(df_sel,test_code):
    import pandas as pd
    import numpy as np

    df_sel['Date Time'] = pd.to_datetime(df_sel['Date Time'], format='%m/%d/%Y %H:%M:%S.%f')
    # df_sel['Session_No'] = df_sel['Session_ID'].cat.codes

    agg_str = lambda x: ','.join(filter(lambda s: pd.notna(s), x))
    pcnf = lambda x: np.nanpercentile(x, 95)
    pcf = lambda x: np.nanpercentile(x, 5)
    cvah = lambda x: (x > 100000).sum()
    cvbt = lambda x: (x > 20000).sum()

    df_sel["Serving Cell Identity"] = df_sel["Serving Cell Identity"].fillna('').replace([np.inf, -np.inf], '').astype(str)
    df_sel["Serving Cell DL EARFCN"] = df_sel["Serving Cell DL EARFCN"].fillna('').replace([np.inf, -np.inf], '').astype(str)

    agg_list ={
    "Date":['min', 'max'],
    "Time":min,
    "Date Time":['min', 'max'],
    "Latitude":['min', 'max'],
    "Longitude":['min', 'max'],
    "IMEI":agg_str,
    "IMSI":'mean',
    "MCC":'mean',
    "MNC":'mean',
    "SIM Operator":agg_str,
    "Technology_Detail":agg_str,
    "MT_Call_Domain":agg_str,
    "MT_Call_Setup_Time_ms": 'mean',
    "Call_Duration":'mean',
    "Call_Setup_Duration":'mean',
    "Call_Setup_Time_sec":'mean',
    "Voice_Call_Setup_Time_sec":'mean',
    "Voice_Call_Success_Rate":'mean',
    "Voice_Call_Setup_Success_Rate":'mean',
    "Voice_Call_Setup_Failure_Rate":'mean',
    "Voice_Call_Completion_Rate":'mean',
    "Voice_Dropped_Call_Rate":'mean',
    "Voice_Speech_Quality_On_Call_Basis":'mean',
    "Call Drop Rate":'mean',
    "Call Dropped Count":'mean',
    "Call Dropped VoLTE Count":'mean',
    "Call Dropped VoNR Count":'mean',
    "Call Block Rate":'mean',
    "Call Blocked Count":'mean',
    "Call Blocked VoLTE Count":'mean',
    "Call Blocked VoNR Count":'mean',
    "Call Success Count":'mean',
    "Call Success Rate":'mean',
    "Call Success VoLTE Count":'mean',
    "Call Success VoNR Count":'mean',
    "Data Radio Bearer":agg_str,
    "Multi RAT Connectivity Mode":agg_str,
    "RadioAccessTechnologyState":'mean',
    "Serving Cell DL EARFCN":agg_str,
    "Serving Cell eMTC Cell Indication":'mean',
    "Serving Cell Frequency Band": agg_str,
    "Serving Cell Identity":agg_str,
    "Serving Cell RSRP (dBm)":'mean',
    "Serving Cell RSRQ (dB)":'mean',
    "Serving Cell RS SINR (dB)":'mean',
    "Serving Cell Serving Cell EN-DC Support":'mean',
    "CallDroppedRate":'mean',
    "DRBDroppedRate":'mean',
    "SipByeReasonPhrase": agg_str,
    "SpeechPathDelayRTT":'mean',
    "SpeechPathDelayOneWay":'mean',
    "RTPInterruptionTimeAudio":'mean',
    "AQMScoreDL":'mean',
    "AQMAlgorithmDL":agg_str,
    "AQM_RAT":agg_str,
    "AQM Algorithm DL":agg_str,
    "AQM Audio Channel Type DL":'mean',
    "AQM Score":'mean',
    "AQM Score DL":'mean',
    "MOS Calculator Version Downlink":agg_str,
    # "Speech Codec":agg_str,
    "AMR_Codec_DL":agg_str,
    "SQ_MOS":'mean',
    "RTP Jitter Audio RFC3550 (ms)":'mean',
    "RTP Lost Packets Rate Audio (%)":'mean',
    "RTP Relative Packet Delay Audio (ms)":'mean',
    "Application Layer kBytes Received (Cumulative)":'mean',
    "Application Layer kBytes Received (Current)":'mean',
    "HTTP_Download_Average_Throughput":'mean',
    "HTTP_Download_Service_Average_Throughput":'mean',
    "HTTP_Download_Session_Success_Ratio":'mean',
    "HTTP_Download_Session_Failure_Ratio":'mean',
    "HTTP_Outcome":agg_str,
    "HTTP_URL":agg_str,
    "HTTP_Download_Bearer":agg_str,
    "HTTP Transfer Time (second)":'mean',
    "Distance": sum,
    "POLQA SWB Score DL": 'mean',
    "Call_Status":agg_str
    }

    pivot_result = pd.pivot_table(df_sel, index=["Test_ID"], values=list(agg_list.keys()), aggfunc=agg_list)

    # df_test = df_sel[["Test_ID",'HTTP Download Throughput (kbps)']]
    # pivot_result_2 = pd.pivot_table(df_test, index='Test_ID', values='HTTP Download Throughput (kbps)', aggfunc=pcnf)
    # print(pivot_result_2.head(10))

    pivot_result.columns = [' '.join(map(str, tpl)) for tpl in pivot_result.columns]
    pivot_result.columns = [s.replace(" <lambda>", '') for s in pivot_result.columns]
    pivot_result.columns = [s.replace(" mean", '') for s in pivot_result.columns]
    df_pivot = pivot_result.reset_index()

    # print(df_pivot)

    def clear_duplicate_values_in_cell(cell_value):
        values = cell_value.split(',')
        unique_values = list(set(values))
        return ','.join(unique_values)
    def clear_duplicate_values_in_column(dataframe, column_name):
        df_copy = dataframe.copy()
        df_copy[column_name] = df_copy[column_name].apply(clear_duplicate_values_in_cell)
        return df_copy

    agg_clear_list = [
        "IMEI",
        "SIM Operator",
        "Technology_Detail",
        "MT_Call_Domain",
        "Data Radio Bearer",
        "Multi RAT Connectivity Mode",
        "Serving Cell Frequency Band",
        "SipByeReasonPhrase",
        "AQMAlgorithmDL",
        "AQM_RAT",
        "AQM Algorithm DL",
        "MOS Calculator Version Downlink",
        "AMR_Codec_DL",
        "HTTP_Outcome",
        "HTTP_URL",
        "HTTP_Download_Bearer",
        "Call_Status",
        "Serving Cell DL EARFCN",
        'Serving Cell Identity'
    ]

    for i in agg_clear_list:
        df_pivot = clear_duplicate_values_in_column(df_pivot,i)

    df_pivot["Duration"] = df_pivot["Date Time max"] - df_pivot["Date Time min"]

    df_pivot['Duration_Seconds'] = df_pivot['Duration'].apply(lambda x: pd.Timedelta(x).total_seconds())

    df_pivot[['Month','Day','Year']] = df_pivot['Date min'].str.split('/', n=3, expand=True)

    df_pivot["Hour"] = df_pivot["Time min"].str[:2]
    df_pivot = df_pivot[df_pivot['Test_ID'] != 0]
    df_pivot["Sequenz_ID_per_File"] = df_pivot["Test_ID"]
    # df_pivot["Test_ID"] =  df_pivot['Year'] +"3" +df_pivot['Month'] +  df_pivot['Day'] + df_pivot["Test_ID"].astype(str)
    df_pivot["Test_ID"] =  df_pivot['Year'] +test_code +df_pivot['Month'].str.zfill(2) +df_pivot['Day'].str.zfill(2) + df_pivot["Test_ID"].astype(str).str.zfill(4)

    col_arrange = [
    "Test_ID",
    "Sequenz_ID_per_File",
    "Day",
    "Month",
    "Year",
    "Hour",
    "Duration_Seconds",
    "IMEI",
    "IMSI",
    "MCC",
    "MNC",
    "Multi RAT Connectivity Mode",
    "RadioAccessTechnologyState",
    "SIM Operator",
    "Distance sum",
    "SipByeReasonPhrase",
    "Technology_Detail",
    "MT_Call_Domain",
    "MT_Call_Setup_Time_ms",
    "MOS Calculator Version Downlink",
    "SpeechPathDelayOneWay",
    "SpeechPathDelayRTT",
    "Call Block Rate",
    "Call Blocked Count",
    "Call Blocked VoLTE Count",
    "Call Blocked VoNR Count",
    "Call Drop Rate",
    "Call Dropped Count",
    "Call Dropped VoLTE Count",
    "Call Dropped VoNR Count",
    "Call Success Count",
    "Call Success Rate",
    "Call Success VoLTE Count",
    "Call Success VoNR Count",
    "CallDroppedRate",
    "Call_Duration",
    "Call_Setup_Duration",
    "Call_Setup_Time_sec",
    "Call_Status",
    "Voice_Call_Completion_Rate",
    "Voice_Call_Setup_Failure_Rate",
    "Voice_Call_Setup_Success_Rate",
    "Voice_Call_Setup_Time_sec",
    "Voice_Call_Success_Rate",
    "Voice_Dropped_Call_Rate",
    "Voice_Speech_Quality_On_Call_Basis",
    "SQ_MOS",
    "POLQA SWB Score DL",
    "AQM Algorithm DL",
    "AQM Audio Channel Type DL",
    "AQM Score",
    "AQM Score DL",
    "AQMAlgorithmDL",
    "AQMScoreDL",
    "AQM_RAT",
    "RTP Jitter Audio RFC3550 (ms)",
    "RTP Lost Packets Rate Audio (%)",
    "RTP Relative Packet Delay Audio (ms)",
    "RTPInterruptionTimeAudio",
    "HTTP_Download_Bearer",
    "HTTP_Outcome",
    "HTTP_URL",
    "DRBDroppedRate",
    "Data Radio Bearer",
    "AMR_Codec_DL",
    "Serving Cell DL EARFCN",
    "Serving Cell Frequency Band",
    "Serving Cell Identity",
    "Serving Cell RS SINR (dB)",
    "Serving Cell RSRP (dBm)",
    "Serving Cell RSRQ (dB)",
    "Serving Cell Serving Cell EN-DC Support",
    "Serving Cell eMTC Cell Indication",
    ]

    df_pivot['Serving Cell Identity'] = "'"  + df_pivot['Serving Cell Identity'].str.lstrip(',').str.split(',').apply(lambda x: ','.join([str(int(float(x[0])))] + [str(int(float(value))) if '.' in value else str(int(value)) for value in x[1:]]))+ "'"
    df_pivot["Serving Cell DL EARFCN"] = "'" + df_pivot['Serving Cell DL EARFCN'].str.lstrip(',').str.split(',').apply(lambda x: ','.join([str(int(float(x[0])))] + [str(int(float(value))) if '.' in value else str(int(value)) for value in x[1:]]))+ "'"

    df_ordered = df_pivot[col_arrange]

    return df_ordered


def excel_report(df_total_stats, merged_df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Font
    import pandas as pd

    Name_list = ["FDFS_DL_Attempts", "FDFS_DL_Success", "FDFS_DL_Failure", "FDFS_DL_Success_Ratio", "FDFS_UL_Attempts", "FDFS_UL_Success", "FDFS_UL_Failure", "FDFS_UL_Success_Ratio", "HTTP_Browsing_Attempts", "HTTP_Browsing_Access_Failure", "HTTP_Browsing_Access_Success", "HTTP_Browsing_Min_Transfer_Time_s", "HTTP_Browsing_10_PCTL_Transfer_Time_s", "HTTP_Browsing_Average_Transfer_Time_s", "HTTP_Browsing_Median_Transfer_Time_s", "HTTP_Browsing_90_PCTL_Transfer_Time_s",
        "HTTP_Browsing_Max_Transfer_Time_s", "Video_Stream_Attempts", "Video_Stream_Failures", "Video_Stream_Access_Time_sec", "Video_Stream_Success_Ratio_PC", "Video_Stream_Reproduction_Start_Delay_sec", "Initial_Latency_Time_sec_Streaming_State_Request_to_Prebuffering_Delay", "Streaming_Reproduction_Cut-off_Ratio", "Streaming_Session_Without_Interruption_Rate_PC", "Streaming_Aggregated_Average_Session_Resolution_p", "Ping_Attempt", "Ping_Success", "Ping_Failures", "Ping_Success_rate"]

    column_headers = [
        "Campaign",
        "Operator",
        "Type_of_Test",
        "Test_Name",
        "Test_Info",
        "Direction",
        "Thread_Info",
        "URL",
        "Size",
        "HTTP_Outcome",
        "ErrorCode_Message",
        "Mean_Data_Rate_Mbit_s",
        "Transfer_Duration_s",
        "TestId",
        "Sessionid",
        "FileID",
        "Sequenz_ID_per_File",
        "File_Name",
        "PCAP_File_Name",
        "Channel",
        "Channel_Description",
        "G_Level_1",
        "G_Level_2",
        "G_Level_3",
        "G_Level_4",
        "G_Level_5",
        "Test_Start_Latitude",
        "Test_Start_Longitude",
        "Test_Distance_m",
        "Year",
        "Quarter",
        "Month",
        "Week",
        "Day",
        "Hour",
        "Test_Start_Time",
        "DNS_1st_Request",
        "DNS_1st_Response",
        "TCP_1st_SYN",
        "TCP_1st_SYN_ACK",
        "Data_1st_recieved",
        "Data_Last_Recieved",
        "Test_End_Time",
        "Test_Duration_s",
        "Transfer_Access_Duration_s",
        "Transfer_Transferred_Bytes",
        "FDTT_Sustainable_Transferred_Bytes",
        "FDTT_Sustainable_Transferred_Time_ms",
        "FDTT_Sustainable_MDR_Mbit_s",
        "http_Browser_Access_Duration_s",
        "http_Browser_Transferred_Bytes",
        "http_Browser_Number_of_Images",
        "http_Browser_1MB_Reached_Duration_ms",
        "Content_Transferred_Bytes",
        "IP_Service_Access_Result",
        "IP_Service_Access_Delay_ms",
        "DNS_Service_Access_Delay_ms",
        "TCP_RTT_Service_Access_Delay_ms",
        "Data_Download_Delay",
        "APN_String",
        "Source_IP",
        "DNS_Server_IPs",
        "Destination_IP",
        "Threads",
        "TCP_Threads_Detailed",
        "DNS_Resolution_Attempts",
        "DNS_Resolution_Success",
        "DNS_Min_Resolution_Time_ms",
        "DNS_Avg_Resolution_Time_ms",
        "DNS_Max_Resolution_Time_ms",
        "DNS_Hosts_Resolved",
        "IP_Layer_Transferred_Bytes_DL",
        "IP_Layer_Transferred_Bytes_UL",
        "PCell_RAT",
        "PCell_RAT_Timeline",
        "TEC_Timeline",
        "TIME_GSM_s",
        "TIME_UMTS_s",
        "TIME_LTE_s",
        "TIME_LTE_CA_s",
        "TIME_NR5G_s",
        "TIME_No_Service_s",
        "IMEI",
        "IMSI",
        "MSISDN",
        "Device",
        "Firmware",
        "Measurement_System",
        "SW_Version",
        "Home_Operator",
        "Home_MCC",
        "Home_MNC",
        "MCC",
        "MNC",
        "CellID",
        "LAC",
        "BCCH",
        "ActiveSet",
        "SC1",
        "BSIC",
        "PCI",
        "LAC_CId_BCCH",
        "MinRSRP",
        "AvgRSRP",
        "MaxRSRP",
        "MinRSRQ",
        "AvgRSRQ",
        "MaxRSRQ",
        "MinSINR",
        "AvgSINR",
        "MaxSINR",
        "MinTxPwr4G",
        "AvgTxPwr4G",
        "MaxTxPwr4G",
        "CA_Configured",
        "CA_Configured_Timeline",
        "CA_Activated",
        "CA_Bands",
        "Test_Bandwith_Average_MHz",
        "PCell_Info_List",
        "PCell_Throughput_Average_Mbit_s",
        "SCell1_Info_List",
        "SCell1_Usage_Ratio",
        "SCell1_Throughput_Average_Mbit_s",
        "SCell2_Info_List",
        "SCell2_Usage_Ratio",
        "SCell2_Throughput_Average_Mbit_s",
        "CQI_LTE_Min",
        "CQI_LTE",
        "CQI_LTE_Max",
        "AvgDLMCS",
        "ShareDLQPSK",
        "ShareDL16QAM",
        "ShareDL64QAM",
        "ShareDL256QAM",
        "DL256QAM_larger_10",
        "PDSCHBytesTransfered",
        "MinNetPDSCHThroughput_Mbit_s",
        "AvgNetPDSCHThroughput_Mbit_s",
        "MaxNetPDSCHThroughput_Mbit_s",
        "ShareULBPSK",
        "ShareULQPSK",
        "ShareUL16QAM",
        "ShareUL64QAM",
        "ShareUL256QAM",
        "UL64QAM_larger_30",
        "PUSCHBytesTransfered",
        "MinNetPUSCHThroughput_Mbit_s",
        "AvgNetPUSCHThroughput_Mbit_s",
        "MaxNetPUSCHThroughput_Mbit_s",
        "NR_RSRP_MIN",
        "NR_RSRP_AVG",
        "NR_RSRP_MAX",
        "NR_RSRQ_MIN",
        "NR_RSRQ_AVG",
        "NR_RSRQ_MAX",
        "NR_SINR_MIN",
        "NR_SINR_AVG",
        "NR_SINR_MAX",
        "NR_TxPwr_MIN",
        "NR_TxPwr_AVG",
        "NR_TxPwr_MAX",
        "LTE_NR_Carriers",
        "LTE_NR_Band_Combination",
        "NrDl_ARFCN",
        "NrDl_PointA",
        "NrDl_PCI",
        "NrDl_Band",
        "NrDl_Total_Time",
        "NrB3_Total_Time",
        "NrDl_Scheduled_PDSCH_Throughput_Avg",
        "NrDl_Net_PDSCH_Throughput_Avg",
        "NrDl_Total_Bandwidth",
        "NrDl_Serving_Beam_Index",
        "NrDl_Numerologies_Number",
        "Numerology1_SCS",
        "Numerology1_BW",
        "NrDl_Initial_BWP_Bandwidth",
        "NrDl_Initial_BWP_SCS",
        "NrDl_BWP1_Bandwidth",
        "NrDl_BWP1_SCS",
        "NrDl_Active_BWP",
        "NrDl_CQI_Avg",
        "NrDl_MCS_Avg",
        "NrDl_QPSK_Ratio",
        "NrDl_16QAM_Ratio",
        "NrDl_64QAM_Ratio",
        "NrDl_256QAM_Ratio",
        "NrDl_MinTBSize",
        "NrDl_AvgTBSize",
        "NrDl_MaxTBSize",
        "NrDl_MinTBRate",
        "NrDl_AvgTBRate",
        "NrDl_MaxTBRate",
        "NrDl_MinRBs",
        "NrDl_AvgRBs",
        "NrDl_MaxRBs",
        "NrDl_Rank",
        "NrDl_Max_MIMO_Layers",
        "NrUl_ARFCN",
        "NrUl_Total_Bandwidth",
        "NrUl_Scheduled_PUSCH_Throughput_Avg_Mbit_s",
        "NrUl_Net_PUSCH_Throughput_Avg_Mbit_s",
        "NrUl_Initial_BWP_Bandwidth",
        "NrUl_Initial_BWP_SCS",
        "NrUl_BWP1_Bandwidth",
        "NrUl_BWP1_SCS",
        "NrUl_Active_BWP",
        "NrUl_MCS_Avg",
        "NrUl_QPSK_Ratio",
        "NrUl_16QAM_Ratio",
        "NrUl_64QAM_Ratio",
        "NrUl_256QAM_Ratio",
        "NrUl_MinTBSize",
        "NrUl_AvgTBSize",
        "NrUl_MaxTBSize",
        "NrUl_MinTBRate",
        "NrUl_AvgTBRate",
        "NrUl_MaxTBRate",
        "NrUl_MinRBs",
        "NrUl_AvgRBs",
        "NrUl_MaxRBs",
        "NrUl_Rank",
        "NrUl_Max_MIMO_Layers",
        "HandoversInfo",
        "Region",
        "Vendor",
        "JOIN_ID",
        "VDF_CELL_NAME",
        "BatteryLevel",
        "BatteryTemp",
        "CPUTemp",
        "CPULoad"
    ]


    df1 = df_total_stats  # pd.DataFrame(data1)
    df2 = pd.DataFrame(columns=column_headers)

    # region creating operator table]

    # df2["Campaign"] = merged_df["Campaign"]
    df2["Operator"] = merged_df["Operator"]
    df2["Type_of_Test"] = merged_df["Type_of_Test"]
    df2["Test_Name"] = merged_df["Test_Name"]
    # df2["Test_Info"] = merged_df[""]
    df2["Direction"] = merged_df["Direction_Test"]
    df2["Thread_Info"] = merged_df["Thread_info"]
    df2["URL"] = merged_df["Grouping_All_URL"]
    # df2["Size"] = merged_df[""]
    df2["HTTP_Outcome"] = merged_df["HTTP_Outcome"]
    # df2["ErrorCode_Message"] = merged_df[""]
    # df2["Mean_Data_Rate_Mbit_s"] = merged_df[""]
    # df2["Transfer_Duration_s"] = merged_df[""]
    df2["TestId"] = merged_df["Test_ID"]
    df2["Sessionid"] = merged_df["Session_No"]
    # df2["FileID"] = merged_df[""]
    df2["Sequenz_ID_per_File"] = merged_df["Sequenz_ID_per_File"]
    # df2["File_Name"] = merged_df[""]
    # df2["PCAP_File_Name"] = merged_df[""]
    # df2["Channel"] = merged_df[""]
    # df2["Channel_Description"] = merged_df[""]
    # df2["G_Level_1"] = merged_df[""]
    # df2["G_Level_2"] = merged_df[""]
    # df2["G_Level_3"] = merged_df[""]
    # df2["G_Level_4"] = merged_df[""]
    # df2["G_Level_5"] = merged_df[""]
    df2["Test_Start_Latitude"] = merged_df["Latitude min"]
    df2["Test_Start_Longitude"] = merged_df["Longitude min"]
    df2["Test_Distance_m"] = merged_df["Distance sum"]
    df2["Year"] = merged_df["Year"]
    # df2["Quarter"] = merged_df[""]
    df2["Month"] = merged_df["Month"]
    # df2["Week"] = merged_df[""]
    df2["Day"] = merged_df["Day"]
    df2["Hour"] = merged_df["Hour"]
    # df2["Test_Start_Time"] = merged_df[""]
    # df2["DNS_1st_Request"] = merged_df[""]
    # df2["DNS_1st_Response"] = merged_df[""]
    # df2["TCP_1st_SYN"] = merged_df[""]
    # df2["TCP_1st_SYN_ACK"] = merged_df[""]
    # df2["Data_1st_recieved"] = merged_df[""]
    # df2["Data_Last_Recieved"] = merged_df[""]
    # df2["Test_End_Time"] = merged_df[""]
    df2["Test_Duration_s"] = merged_df["Duration_Seconds"]
    # df2["Transfer_Access_Duration_s"] = merged_df[""]
    # df2["Transfer_Transferred_Bytes"] = merged_df[""]
    # df2["FDTT_Sustainable_Transferred_Bytes"] = merged_df[""]
    # df2["FDTT_Sustainable_Transferred_Time_ms"] = merged_df[""]
    # df2["FDTT_Sustainable_MDR_Mbit_s"] = merged_df[""]
    # df2["http_Browser_Access_Duration_s"] = merged_df[""]
    # df2["http_Browser_Transferred_Bytes"] = merged_df[""]
    # df2["http_Browser_Number_of_Images"] = merged_df[""]
    # df2["http_Browser_1MB_Reached_Duration_ms"] = merged_df[""]
    # df2["Content_Transferred_Bytes"] = merged_df[""]
    # df2["IP_Service_Access_Result"] = merged_df[""]
    # df2["IP_Service_Access_Delay_ms"] = merged_df[""]
    # df2["DNS_Service_Access_Delay_ms"] = merged_df[""]
    # df2["TCP_RTT_Service_Access_Delay_ms"] = merged_df[""]
    # df2["Data_Download_Delay (SYNACK -> 1stData)"] = merged_df[""]
    # df2["APN_String"] = merged_df[""]
    # df2["Source_IP"] = merged_df[""]
    # df2["DNS_Server_IPs"] = merged_df[""]
    # df2["Destination_IP"] = merged_df[""]
    # df2["Threads"] = merged_df[""]
    # df2["TCP_Threads_Detailed"] = merged_df[""]
    # df2["DNS_Resolution_Attempts"] = merged_df[""]
    # df2["DNS_Resolution_Success"] = merged_df[""]
    # df2["DNS_Min_Resolution_Time_ms"] = merged_df[""]
    # df2["DNS_Avg_Resolution_Time_ms"] = merged_df[""]
    # df2["DNS_Max_Resolution_Time_ms"] = merged_df[""]
    # df2["DNS_Hosts_Resolved"] = merged_df[""]
    # df2["IP_Layer_Transferred_Bytes_DL"] = merged_df[""]
    # df2["IP_Layer_Transferred_Bytes_UL"] = merged_df[""]
    # df2["PCell_RAT"] = merged_df["PCell_RAT"]
    # df2["PCell_RAT_Timeline"] = merged_df[""]
    # df2["TEC_Timeline"] = merged_df[""]
    # df2["TIME_GSM_s"] = merged_df[""]
    # df2["TIME_UMTS_s"] = merged_df[""]
    # df2["TIME_LTE_s"] = merged_df[""]
    # df2["TIME_LTE_CA_s"] = merged_df[""]
    # df2["TIME_NR5G_s"] = merged_df[""]
    # df2["TIME_No_Service_s"] = merged_df[""]
    # df2["IMEI"] = merged_df[""]
    # df2["IMSI"] = merged_df[""]
    # df2["MSISDN"] = merged_df[""]
    # df2["Device"] = merged_df[""]
    # df2["Firmware"] = merged_df[""]
    # df2["Measurement_System"] = merged_df[""]
    # df2["SW_Version"] = merged_df[""]
    # df2["Home_Operator"] = merged_df[""]
    # df2["Home_MCC"] = merged_df[""]
    # df2["Home_MNC"] = merged_df[""]
    # df2["MCC"] = merged_df[""]
    # df2["MNC"] = merged_df[""]
    # df2["CellID"] = merged_df[""]
    # df2["LAC"] = merged_df[""]
    # df2["BCCH"] = merged_df[""]
    # df2["ActiveSet"] = merged_df[""]
    # df2["SC1"] = merged_df[""]
    # df2["BSIC"] = merged_df[""]
    # df2["PCI"] = merged_df[""]
    # df2["LAC_CId_BCCH"] = merged_df[""]
    # df2["MinRSRP"] = merged_df["Serving Cell RSRP (dBm) min"]
    # df2["AvgRSRP"] = merged_df["Serving Cell RSRP (dBm)"]
    # df2["MaxRSRP"] = merged_df["Serving Cell RSRP (dBm) max"]
    # df2["MinRSRQ"] = merged_df["Serving Cell RSRQ (dB) min"]
    # df2["AvgRSRQ"] = merged_df["Serving Cell RSRQ (dB)"]
    # df2["MaxRSRQ"] = merged_df["Serving Cell RSRQ (dB) max"]
    # df2["MinSINR"] = merged_df["Serving Cell RS SINR (dB) min"]
    # df2["AvgSINR"] = merged_df["Serving Cell RS SINR (dB)"]
    # df2["MaxSINR"] = merged_df["Serving Cell RS SINR (dB) max"]
    # df2["MinTxPwr4G"] = merged_df[""]
    # df2["AvgTxPwr4G"] = merged_df[""]
    # df2["MaxTxPwr4G"] = merged_df[""]
    # df2["CA_Configured"] = merged_df[""]
    # df2["CA_Configured_Timeline"] = merged_df[""]
    # df2["CA_Activated"] = merged_df[""]
    # df2["CA_Bands"] = merged_df[""]
    # df2["Test_Bandwith_Average_MHz"] = merged_df[""]
    # df2["PCell_Info_List"] = merged_df[""]
    # df2["PCell_Throughput_Average_Mbit_s"] = merged_df[""]
    # df2["SCell1_Info_List"] = merged_df[""]
    # df2["SCell1_Usage_Ratio"] = merged_df[""]
    # df2["SCell1_Throughput_Average_Mbit_s"] = merged_df[""]
    # df2["SCell2_Info_List"] = merged_df[""]
    # df2["SCell2_Usage_Ratio"] = merged_df[""]
    # df2["SCell2_Throughput_Average_Mbit_s"] = merged_df[""]
    # df2["CQI_LTE_Min"] = merged_df[""]
    # df2["CQI_LTE"] = merged_df[""]
    # df2["CQI_LTE_Max"] = merged_df[""]
    # df2["AvgDLMCS"] = merged_df[""]
    # df2["ShareDLQPSK"] = merged_df[""]
    # df2["ShareDL16QAM"] = merged_df[""]
    # df2["ShareDL64QAM"] = merged_df[""]
    # df2["ShareDL256QAM"] = merged_df[""]
    # df2["DL256QAM_larger_10"] = merged_df[""]
    # df2["PDSCHBytesTransfered"] = merged_df[""]
    # df2["MinNetPDSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["AvgNetPDSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["MaxNetPDSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["ShareULBPSK"] = merged_df[""]
    # df2["ShareULQPSK"] = merged_df[""]
    # df2["ShareUL16QAM"] = merged_df[""]
    # df2["ShareUL64QAM"] = merged_df[""]
    # df2["ShareUL256QAM"] = merged_df[""]
    # df2["UL64QAM_larger_30"] = merged_df[""]
    # df2["PUSCHBytesTransfered"] = merged_df[""]
    # df2["MinNetPUSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["AvgNetPUSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["MaxNetPUSCHThroughput_Mbit_s"] = merged_df[""]
    # df2["NR_RSRP_MIN"] = merged_df[""]
    # df2["NR_RSRP_AVG"] = merged_df[""]
    # df2["NR_RSRP_MAX"] = merged_df[""]
    # df2["NR_RSRQ_MIN"] = merged_df[""]
    # df2["NR_RSRQ_AVG"] = merged_df[""]
    # df2["NR_RSRQ_MAX"] = merged_df[""]
    # df2["NR_SINR_MIN"] = merged_df[""]
    # df2["NR_SINR_AVG"] = merged_df[""]
    # df2["NR_SINR_MAX"] = merged_df[""]
    # df2["NR_TxPwr_MIN"] = merged_df[""]
    # df2["NR_TxPwr_AVG"] = merged_df[""]
    # df2["NR_TxPwr_MAX"] = merged_df[""]
    # df2["LTE_NR_Carriers"] = merged_df[""]
    # df2["LTE_NR_Band_Combination"] = merged_df[""]
    # df2["NrDl_ARFCN"] = merged_df[""]
    # df2["NrDl_PointA"] = merged_df[""]
    # df2["NrDl_PCI"] = merged_df[""]
    # df2["NrDl_Band"] = merged_df[""]
    # df2["NrDl_Total_Time"] = merged_df[""]
    # df2["NrB3_Total_Time"] = merged_df[""]
    # df2["NrDl_Scheduled_PDSCH_Throughput_Avg"] = merged_df[""]
    # df2["NrDl_Net_PDSCH_Throughput_Avg"] = merged_df[""]
    # df2["NrDl_Total_Bandwidth"] = merged_df[""]
    # df2["NrDl_Serving_Beam_Index"] = merged_df[""]
    # df2["NrDl_Numerologies_Number"] = merged_df[""]
    # df2["Numerology1_SCS"] = merged_df[""]
    # df2["Numerology1_BW"] = merged_df[""]
    # df2["NrDl_Initial_BWP_Bandwidth"] = merged_df[""]
    # df2["NrDl_Initial_BWP_SCS"] = merged_df[""]
    # df2["NrDl_BWP1_Bandwidth"] = merged_df[""]
    # df2["NrDl_BWP1_SCS"] = merged_df[""]
    # df2["NrDl_Active_BWP"] = merged_df[""]
    # df2["NrDl_CQI_Avg"] = merged_df[""]
    # df2["NrDl_MCS_Avg"] = merged_df[""]
    # df2["NrDl_QPSK_Ratio"] = merged_df[""]
    # df2["NrDl_16QAM_Ratio"] = merged_df[""]
    # df2["NrDl_64QAM_Ratio"] = merged_df[""]
    # df2["NrDl_256QAM_Ratio"] = merged_df[""]
    # df2["NrDl_MinTBSize"] = merged_df[""]
    # df2["NrDl_AvgTBSize"] = merged_df[""]
    # df2["NrDl_MaxTBSize"] = merged_df[""]
    # df2["NrDl_MinTBRate"] = merged_df[""]
    # df2["NrDl_AvgTBRate"] = merged_df[""]
    # df2["NrDl_MaxTBRate"] = merged_df[""]
    # df2["NrDl_MinRBs"] = merged_df[""]
    # df2["NrDl_AvgRBs"] = merged_df[""]
    # df2["NrDl_MaxRBs"] = merged_df[""]
    # df2["NrDl_Rank"] = merged_df[""]
    # df2["NrDl_Max_MIMO_Layers"] = merged_df[""]
    # df2["NrUl_ARFCN"] = merged_df[""]
    # df2["NrUl_Total_Bandwidth"] = merged_df[""]
    # df2["NrUl_Scheduled_PUSCH_Throughput_Avg_Mbit_s"] = merged_df[""]
    # df2["NrUl_Net_PUSCH_Throughput_Avg_Mbit_s"] = merged_df[""]
    # df2["NrUl_Initial_BWP_Bandwidth"] = merged_df[""]
    # df2["NrUl_Initial_BWP_SCS"] = merged_df[""]
    # df2["NrUl_BWP1_Bandwidth"] = merged_df[""]
    # df2["NrUl_BWP1_SCS"] = merged_df[""]
    # df2["NrUl_Active_BWP"] = merged_df[""]
    # df2["NrUl_MCS_Avg"] = merged_df[""]
    # df2["NrUl_QPSK_Ratio"] = merged_df[""]
    # df2["NrUl_16QAM_Ratio"] = merged_df[""]
    # df2["NrUl_64QAM_Ratio"] = merged_df[""]
    # df2["NrUl_256QAM_Ratio"] = merged_df[""]
    # df2["NrUl_MinTBSize"] = merged_df[""]
    # df2["NrUl_AvgTBSize"] = merged_df[""]
    # df2["NrUl_MaxTBSize"] = merged_df[""]
    # df2["NrUl_MinTBRate"] = merged_df[""]
    # df2["NrUl_AvgTBRate"] = merged_df[""]
    # df2["NrUl_MaxTBRate"] = merged_df[""]
    # df2["NrUl_MinRBs"] = merged_df[""]
    # df2["NrUl_AvgRBs"] = merged_df[""]
    # df2["NrUl_MaxRBs"] = merged_df[""]
    # df2["NrUl_Rank"] = merged_df[""]
    # df2["NrUl_Max_MIMO_Layers"] = merged_df[""]
    # df2["HandoversInfo"] = merged_df[""]
    # df2["Region"] = merged_df[""]
    # df2["Vendor"] = merged_df[""]
    # df2["JOIN_ID"] = merged_df[""]
    # df2["VDF_CELL_NAME"] = merged_df[""]
    # df2["BatteryLevel"] = merged_df[""]
    # df2["BatteryTemp"] = merged_df[""]
    # df2["CPUTemp"] = merged_df[""]
    # df2["CPULoad"] = merged_df[""]
    df2["RadioAccessTechnologyState"] = merged_df["RadioAccessTechnologyState"]
    mapping = {2: 'GSM/WCDMA', 3: 'LTE', 7: 'EN-DC'}
    df2['RadioAccessTechnology'] = df2["RadioAccessTechnologyState"].map(mapping).fillna('LTE/EN-DC')
    df2["HTTP_Upload_Session_Failure_Ratio"] = merged_df["HTTP_Upload_Session_Failure_Ratio"]
    df2["HTTP_Upload_Session_Success_Ratio"] = merged_df["HTTP_Upload_Session_Success_Ratio"]
    df2["HTTP Download Throughput (kbps) pctl95"] = merged_df["HTTP Download Throughput (kbps) <lambda_0>"]
    df2["HTTP Download Throughput (kbps) pctl5"] = merged_df["HTTP Download Throughput (kbps) <lambda_1>"]
    df2["HTTP Download Throughput (kbps) >100"] = merged_df["HTTP Download Throughput (kbps) <lambda_2>"]
    df2["HTTP Download Throughput (kbps) >20"] = merged_df["HTTP Download Throughput (kbps) <lambda_3>"]
    df2["HTTP Download Throughput (kbps) max"] = merged_df["HTTP Download Throughput (kbps) max"]
    df2["HTTP Download Throughput (kbps)"] = merged_df["HTTP Download Throughput (kbps)"]
    df2["HTTP Download Throughput (kbps) median"] = merged_df["HTTP Download Throughput (kbps) median"]
    df2["HTTP Download Throughput (kbps) min"] = merged_df["HTTP Download Throughput (kbps) min"]
    df2["HTTP_Upload_Average_Throughput"] = merged_df["HTTP_Upload_Average_Throughput"]
    df2["Streaming_Video_Session_Failure_Ratio"]=merged_df["Streaming_Video_Session_Failure_Ratio"]
    df2["Streaming_Video_Session_Cutoff_Ratio"]=merged_df["Streaming_Video_Session_Cutoff_Ratio"]
    df2["Streaming_Video_Session_Success_Ratio"]=merged_df["Streaming_Video_Session_Success_Ratio"]
    df2["Streaming_Average_Throughput"]=merged_df["Streaming_Average_Throughput"]
    df2["Streaming_Duration"]=merged_df["Streaming_Duration"]
    df2["Streaming_Service_Access_Time_sec"] = merged_df["Streaming_Service_Access_Time_sec"]
    df2["Streaming_Reproduction_Start_Delay_sec"] = merged_df["Streaming_Reproduction_Start_Delay_sec"]
    df2["Streaming_State_Request_to_Prebuffering_Delay"] = merged_df["Streaming_State_Request_to_Prebuffering_Delay"]
    df2["Streaming_Impairment_Free_Video_Session_Ratio"] = merged_df["Streaming_Impairment_Free_Video_Session_Ratio"]
    df2["Streaming_Reproduction_Cutoff_Ratio"] = merged_df["Streaming_Reproduction_Cutoff_Ratio"]
    df2["Streaming_Session_Without_Interruption_Rate"] = merged_df["Streaming_Session_Without_Interruption_Rate"]
    df2["Streaming_Average_Session_Resolution		"] = merged_df["Streaming_Average_Session_Resolution"]
    df2["Ping_Count_Attempts "] = merged_df["Ping_Count_Attempts"]
    df2["Ping_Count_Failed "] = merged_df["Ping_Count_Failed"]
    df2["Ping_Count_Success"] = merged_df["Ping_Count_Success"]
    df2["Ping_Packet_Loss_Rate"] = merged_df["Ping_Packet_Loss_Rate"]
    df2["Ping_Roundtrip_Time_sec"] = merged_df["Ping_Roundtrip_Time_sec"]

    # endregion

    # Specify the Excel file name
    excel_file = 'Tools/output_file.xlsx'

    # Create an Excel writer
    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        # Write each DataFrame to a different sheet
        df1.to_excel(writer, sheet_name='Master', index=False)
        df2.to_excel(writer, sheet_name='Vodafone', index=False)

    wb = load_workbook(excel_file)

    sheet = wb["Master"]
    # Adjust the width of the first column
    sheet.column_dimensions['A'].width = 60  # Set the width to 15

    sheet = wb["Vodafone"]
    sheet.column_dimensions['A'].width = 30  # Set the width to 15

    # Save the changes to the Excel file
    wb.save(excel_file)
    print(f'Excel file "{excel_file}" created and styled successfully.')
    return df2
def add_dataframe_excel(input_file,output_file,sheet_name,df):
    import openpyxl
    import pandas as pd
    workbook = openpyxl.load_workbook(input_file, keep_vba=True)
    worksheet = workbook[sheet_name]

    df['Test_ID'] = pd.to_numeric(df['Test_ID'])

    for row in df.iterrows():
        worksheet.append(row[1].tolist())
    workbook.save(output_file)